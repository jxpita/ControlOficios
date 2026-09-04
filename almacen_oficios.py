"""
Capa de almacenamiento de OFICIOS.

*** Punto clave de arquitectura ***
Toda la app (interfaz y métricas) habla SOLO con las funciones de este módulo.
El día que migres a SQLite o a un motor SQL (ver README), reescribes el
cuerpo de estas funciones y NO tocas la interfaz ni las métricas.

Formato: oficios.dat cifrado con Fernet; internamente una lista JSON.
Cada oficio recibe una Referencia UDC generada por el sistema:
    REQ-UDC-<sigla>-<año>-NNNN    REQ-UDC-SB-2026-0001, REQ-UDC-FGE-2026-0001, ...
La sigla la aporta la institución que remite el oficio; el secuencial es
independiente para cada una y se reinicia cada año (ver el módulo parametros).
"""
import csv
import json
import shutil
from datetime import datetime, date
from pathlib import Path
from typing import List, Dict, Optional, Tuple

from cryptography.fernet import InvalidToken

from configuracion import (
    ARCHIVO_OFICIOS, PREFIJO_REFERENCIA, ESTADOS, ROLES_GESTORES, DIR_RESPUESTAS,
    DIR_DOCUMENTOS, EXTENSIONES_DOCUMENTO, ROL_ADMINISTRADOR, ROL_SUPERUSUARIO,
    TIPOS_IDENTIFICACION, TIPOS_IMPLICADO, VALORES_LCI, PRIORIDADES,
)
from cifrado import cifrar, descifrar
import registro_actividad
import permisos
import parametros
import bloqueo


# --- Persistencia ------------------------------------------------------------
def _leer_registros() -> List[Dict]:
    if not ARCHIVO_OFICIOS.exists():
        return []
    try:
        return json.loads(descifrar(ARCHIVO_OFICIOS.read_bytes()))
    except InvalidToken:
        raise ValueError(
            "El archivo de oficios fue alterado o la clave no coincide."
        )


def _guardar_registros(registros: List[Dict]) -> None:
    permisos.escribir_bytes_protegido(
        ARCHIVO_OFICIOS,
        cifrar(json.dumps(registros, ensure_ascii=False, indent=2)),
    )


# --- Validaciones y referencia ----------------------------------------------
def _validar_fecha(texto: str, campo: str) -> str:
    """Valida el formato y que la fecha no sea futura: no se puede registrar
    algo que todavía no ha ocurrido."""
    try:
        valor = datetime.strptime(texto, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        raise ValueError(f"{campo} debe tener formato AAAA-MM-DD.")
    if valor > date.today():
        raise ValueError(f"{campo} no puede ser posterior a hoy.")
    return texto


def _generar_referencia(registros: List[Dict], institucion: str) -> str:
    """Genera la Referencia UDC:  REQ-UDC-<sigla>-<año>-<secuencial de 4 dígitos>.

    El secuencial es **independiente para cada institución** y **se reinicia
    cada año**: REQ-UDC-SB-2026-0001, REQ-UDC-SB-2026-0002… y en paralelo
    REQ-UDC-FGE-2026-0001, REQ-UDC-FGE-2026-0002…; al cambiar de año ambas
    vuelven a 0001.

    Arranca desde el último número usado fuera del sistema, que un gestor puede
    configurar por institución (ver `parametros`). Se usa
    max(configurado, mayor existente) + 1, de modo que nunca se genera una
    referencia duplicada aunque se reconfigure el valor inicial o queden huecos.
    """
    sigla = parametros.sigla_de(institucion)
    anio = parametros.anio_vigente()
    prefijo = f"{PREFIJO_REFERENCIA}-{sigla}-{anio}-"
    secuencial_max = parametros.obtener_secuencial_inicial(institucion, anio)
    for registro in registros:
        referencia = (registro.get("referencia", "") or "").upper()
        if referencia.startswith(prefijo.upper()):
            try:
                secuencial_max = max(secuencial_max, int(referencia.rsplit("-", 1)[1]))
            except ValueError:
                pass
    return f"{prefijo}{secuencial_max + 1:04d}"               # primero -> 0001


# --- Reglas de negocio: relación responsable / estado -----------------------
def _resolver_estado(nombre_empleado: str, estado: str,
                     fecha_respuesta: str = "") -> str:
    """Aplica las reglas de negocio entre responsable, fecha de respuesta y
    estado.

    - Con fecha de respuesta: el oficio ya fue respondido, así que el estado es
      siempre "Finalizado" (y exige responsable).
    - Sin responsable: el único estado válido es "Por asignar". Si se pidió
      "En proceso" o "Finalizado" se lanza un error (esos estados exigen
      responsable).
    - Con responsable: no puede quedar en "Por asignar"; al asignar un
      responsable el oficio pasa automáticamente a "En proceso". Si el estado
      pedido es "En proceso" o "Finalizado" se respeta.
    Devuelve el estado ya corregido.
    """
    tiene_responsable = bool((nombre_empleado or "").strip())
    if (fecha_respuesta or "").strip():
        # Si ya hay respuesta, el oficio está finalizado.
        if not tiene_responsable:
            raise ValueError(
                "Para registrar una fecha de respuesta debe asignar un "
                "responsable (el oficio pasa a \"Finalizado\")."
            )
        return "Finalizado"
    if not tiene_responsable:
        if estado in ("En proceso", "Finalizado"):
            raise ValueError(
                f"El estado \"{estado}\" requiere un responsable asignado."
            )
        return "Por asignar"
    # Con responsable asignado
    if estado == "Por asignar":
        return "En proceso"
    return estado


def _rol_de(usuario: str) -> str:
    """Rol del usuario indicado, o '' si no existe."""
    if not (usuario or "").strip():
        return ""
    # Import diferido: `autenticacion` no depende de este módulo, pero así se
    # mantiene la dependencia acotada a esta comprobación.
    import autenticacion
    objetivo = (usuario or "").strip().lower()
    for registro in autenticacion.listar_usuarios():
        if registro["usuario"].strip().lower() == objetivo:
            return registro.get("rol", "")
    return ""


def _validar_asignacion(id_empleado: str, actor_rol: str) -> None:
    """Un ADMINISTRADOR no puede asignar oficios a un superusuario.

    El superusuario sí puede asignar a cualquiera. Se comprueba aquí, en el
    almacén, para que la regla se cumpla venga de donde venga la llamada.
    """
    if actor_rol != ROL_ADMINISTRADOR:
        return
    if _rol_de(id_empleado) == ROL_SUPERUSUARIO:
        raise ValueError(
            "Como administrador no puede asignar oficios a un superusuario. "
            "Esa asignación corresponde a un superusuario."
        )


def _validar_tipo_accion(tipo: str) -> str:
    """Comprueba el tipo de acción contra el catálogo mantenible.

    Import diferido: `tipos_accion` consulta este módulo para saber cuáles
    están en uso, así que la dependencia se resuelve en el momento de la
    llamada y no al importar.
    """
    import tipos_accion
    return tipos_accion.validar(tipo)


def _validar_prioridad(prioridad: str) -> str:
    """Prioridad de atención del oficio. Es opcional: el histórico que se carga
    desde la matriz puede no traerla."""
    prioridad = " ".join(str(prioridad or "").split()).capitalize()
    if prioridad and prioridad not in PRIORIDADES:
        raise ValueError(
            f"La prioridad «{prioridad}» no es válida. "
            f"Opciones: {', '.join(PRIORIDADES)}."
        )
    return prioridad


def _validar_cantidad_investigados(valor) -> str:
    """Cantidad de investigados: número entero no negativo, opcional.
    Devuelve el valor normalizado como texto ('' si no se indicó)."""
    texto = str(valor if valor is not None else "").strip()
    if not texto:
        return ""
    try:
        cantidad = int(texto)
    except ValueError:
        raise ValueError("La cantidad de investigados debe ser un número entero.")
    if cantidad < 0:
        raise ValueError("La cantidad de investigados no puede ser negativa.")
    return str(cantidad)


def _validar_fecha_asignacion(fecha_asignacion: str, fecha_recepcion: str) -> str:
    """Fecha de asignación (opcional). No puede ser anterior a la de recepción:
    un oficio se asigna después de recibirlo. Devuelve la fecha normalizada."""
    fecha_asignacion = (fecha_asignacion or "").strip()
    if not fecha_asignacion:
        return ""
    _validar_fecha(fecha_asignacion, "Fecha de asignación")
    if (datetime.strptime(fecha_asignacion, "%Y-%m-%d")
            < datetime.strptime(fecha_recepcion, "%Y-%m-%d")):
        raise ValueError(
            "La fecha de asignación no puede ser anterior a la fecha de recepción."
        )
    return fecha_asignacion


def _exigir_cantidad_coherente(registro: Dict, cantidad: str) -> None:
    """La cantidad de investigados no se teclea si el oficio tiene detalle.

    Con implicados anotados el número lo cuenta la lista (ver
    `_sincronizar_investigados`); dejar que además se escriba a mano solo sirve
    para que digan cosas distintas.
    """
    implicados = registro.get("implicados") or []
    if implicados and cantidad and cantidad != str(len(implicados)):
        raise ValueError(
            f"La cantidad de investigados la cuentan los implicados del "
            f"oficio ({len(implicados)}). Añada o quite personas en el "
            f"detalle del oficio (doble clic) en vez de escribir el número."
        )


def _exigir_no_anulado(registro: Dict) -> None:
    """Un oficio anulado está retirado de la operación: no admite cambios de
    trámite hasta que se reactive."""
    if registro.get("anulado"):
        raise ValueError(
            "Este oficio está anulado. Reactívelo antes de modificarlo."
        )


def _exigir_datos_para_finalizar(estado: str, archivo_respuesta: str,
                                 fecha_asignacion: str, fecha_respuesta: str,
                                 estado_previo: str = "") -> None:
    """Un oficio solo puede PASAR a "Finalizado" si el expediente está completo:
    fecha de asignación, fecha de respuesta y la respuesta en PDF adjunta.

    La regla se aplica al MARCARLO como finalizado. Los que ya estaban
    finalizados (por ser anteriores a estas exigencias o por venir de una carga
    masiva de histórico) siguen siendo editables —por ejemplo, para corregirles
    la observación—, porque de lo contrario quedarían bloqueados para siempre.
    """
    if estado != "Finalizado" or estado_previo == "Finalizado":
        return
    faltan = []
    if not (fecha_asignacion or "").strip():
        faltan.append("la fecha de asignación")
    if not (fecha_respuesta or "").strip():
        faltan.append("la fecha de respuesta")
    if not (archivo_respuesta or "").strip():
        faltan.append("la respuesta en PDF adjunta")
    if faltan:
        if len(faltan) > 1:
            detalle = ", ".join(faltan[:-1]) + " y " + faltan[-1]
        else:
            detalle = faltan[0]
        raise ValueError(f"Para finalizar el oficio falta {detalle}.")


def _guardar_documento(referencia: str, ruta_origen: str, carpeta: Path,
                       extensiones, etiqueta: str) -> str:
    """Copia un adjunto a la carpeta de datos y devuelve su nombre de archivo.

    El archivo se guarda como '<referencia><extensión original>', de modo que
    cada oficio tiene como mucho un documento de cada tipo.
    """
    origen = Path(ruta_origen)
    if not origen.exists():
        raise ValueError("No se encontró el archivo seleccionado.")
    extension = origen.suffix.lower()
    if extension not in extensiones:
        admitidas = " o ".join(e.upper().lstrip(".") for e in extensiones)
        raise ValueError(f"{etiqueta} debe ser un archivo {admitidas}.")
    nombre = f"{referencia}{extension}"
    destino = carpeta / nombre
    try:
        # El destino puede existir en solo lectura de una carga previa.
        permisos.hacer_escribible(destino)
        shutil.copyfile(origen, destino)
        permisos.proteger(destino)
    except OSError as error:
        raise ValueError(f"No se pudo guardar el archivo: {error}")
    return nombre


# --- Operaciones -------------------------------------------------------------
def _validar_fecha_respuesta(fecha_respuesta: str, fecha_recepcion: str) -> str:
    """Valida la fecha de respuesta (opcional). No puede ser anterior a la de
    recepción: no se responde un oficio antes de recibirlo. Devuelve la fecha
    normalizada ('' si no se indicó)."""
    fecha_respuesta = (fecha_respuesta or "").strip()
    if not fecha_respuesta:
        return ""
    _validar_fecha(fecha_respuesta, "Fecha de respuesta")
    if (datetime.strptime(fecha_respuesta, "%Y-%m-%d")
            < datetime.strptime(fecha_recepcion, "%Y-%m-%d")):
        raise ValueError(
            "La fecha de respuesta no puede ser anterior a la fecha de recepción."
        )
    return fecha_respuesta


@bloqueo.con_bloqueo("oficios")
def registrar_oficio(codigo_oficio: str, fecha_recepcion: str, fecha_oficio: str,
                     id_empleado: str, nombre_empleado: str, estado: str,
                     registrado_por: str, fecha_respuesta: str = "",
                     observacion: str = "", causal_oficio: str = "",
                     actor_rol: str = None,
                     ruta_documento: str = "", fecha_asignacion: str = "",
                     cantidad_investigados="", ruta_respuesta: str = "",
                     institucion: str = "", tipo_accion: str = "",
                     prioridad: str = "",
                     implicados: Optional[List[Dict]] = None) -> str:
    """`codigo_oficio` es la **Referencia oficio** (obligatoria).
    `causal_oficio` es opcional.

    `institucion` es la entidad que remite el oficio y decide la nomenclatura
    de la Referencia UDC (REQ-UDC-SB-… o REQ-UDC-FGE-…); es **obligatoria**.
    `tipo_accion` es lo que el oficio solicita y debe estar en el catálogo.

    `ruta_documento` es el documento del oficio (PDF o Word) y es
    **obligatorio**: no se registra un oficio sin su soporte.

    `ruta_respuesta` es opcional y solo hace falta para registrar de entrada un
    oficio ya finalizado, porque finalizar exige tener la respuesta adjunta.

    Un **usuario regular** solo puede registrar oficios **auto-asignados**: el
    responsable debe ser él mismo. Asignar a otra persona queda reservado a
    superusuario y administradores, y un administrador no puede asignárselos a
    un superusuario.
    """
    codigo_oficio = codigo_oficio.strip()
    if not codigo_oficio:
        raise ValueError("Debe ingresar la referencia del oficio o circular.")
    causal_oficio = (causal_oficio or "").strip()
    if not (ruta_documento or "").strip():
        raise ValueError(
            "Debe adjuntar el documento del oficio en formato PDF o Word (.docx)."
        )
    # La institución fija la nomenclatura de la referencia, así que se valida
    # antes que nada.
    institucion = parametros.validar_institucion(institucion)
    tipo_accion = _validar_tipo_accion(tipo_accion)

    # Auto-asignación obligatoria para los usuarios regulares.
    if actor_rol is not None and actor_rol not in ROLES_GESTORES:
        if (id_empleado or "").strip().lower() != (registrado_por or "").strip().lower():
            raise ValueError(
                "Solo puede registrar oficios asignados a usted mismo. "
                "Asignarlos a otra persona corresponde a un administrador."
            )
    _validar_asignacion(id_empleado, actor_rol)
    _validar_fecha(fecha_recepcion, "Fecha de recepción")
    _validar_fecha(fecha_oficio, "Fecha de oficio")
    # La fecha de oficio no puede ser posterior a la de recepción: no se puede
    # recibir un oficio antes de que exista.
    if datetime.strptime(fecha_oficio, "%Y-%m-%d") > datetime.strptime(fecha_recepcion, "%Y-%m-%d"):
        raise ValueError(
            "La fecha de oficio no puede ser posterior a la fecha de recepción."
        )
    # Fecha de respuesta, fecha de asignación, cantidad de investigados y
    # observación son opcionales.
    fecha_respuesta = _validar_fecha_respuesta(fecha_respuesta, fecha_recepcion)
    fecha_asignacion = _validar_fecha_asignacion(fecha_asignacion, fecha_recepcion)
    cantidad_investigados = _validar_cantidad_investigados(cantidad_investigados)
    prioridad = _validar_prioridad(prioridad)
    # Personas investigadas anotadas en el propio alta. Si vienen, son ellas
    # las que fijan la cantidad de investigados.
    detalle = []
    for numero, datos in enumerate(implicados or [], start=1):
        persona = validar_implicado(
            datos.get("nombre", ""), datos.get("tipo_identificacion", ""),
            datos.get("identificacion", ""), datos.get("tipo_implicado", ""),
            datos.get("lci", "No"))
        persona["id"] = numero
        detalle.append(persona)
    if detalle:
        cantidad_investigados = str(len(detalle))
    observacion = (observacion or "").strip()
    if estado not in ESTADOS:
        raise ValueError("Estado no válido.")

    # El responsable es opcional. Las reglas ajustan el estado en consecuencia
    # (incluida la fecha de respuesta, que implica "Finalizado").
    nombre_empleado = (nombre_empleado or "").strip()
    id_empleado = (id_empleado or "").strip()
    estado = _resolver_estado(nombre_empleado, estado, fecha_respuesta)
    # Finalizar exige el expediente completo, así que para registrar de entrada
    # un oficio ya finalizado hay que aportarlo todo en el mismo formulario.
    _exigir_datos_para_finalizar(estado, ruta_respuesta, fecha_asignacion,
                                 fecha_respuesta)

    registros = _leer_registros()
    # La referencia del oficio no puede repetirse (aunque la Referencia UDC sea
    # única). Se compara sin distinguir mayúsculas/minúsculas ni espacios.
    codigo_normalizado = codigo_oficio.casefold()
    for registro in registros:
        if esta_anulado(registro):
            continue          # un oficio retirado no reserva su referencia
        if registro.get("codigo_oficio", "").strip().casefold() == codigo_normalizado:
            raise ValueError(
                f"Ya existe un oficio con la referencia \"{codigo_oficio}\". "
                "La referencia del oficio no puede repetirse."
            )
    referencia = _generar_referencia(registros, institucion)
    # Los adjuntos se copian una vez conocida la referencia, que da nombre al
    # archivo. Si algo falla, se lanza el error antes de guardar el registro.
    archivo_oficio = _guardar_documento(
        referencia, ruta_documento, DIR_DOCUMENTOS, EXTENSIONES_DOCUMENTO,
        "El documento del oficio")
    archivo_respuesta = ""
    if (ruta_respuesta or "").strip():
        archivo_respuesta = _guardar_documento(
            referencia, ruta_respuesta, DIR_RESPUESTAS, (".pdf",),
            "La respuesta")
    ahora = datetime.now().isoformat(timespec="seconds")
    registros.append({
        "referencia": referencia,          # Referencia UDC
        "institucion": institucion,        # decide la sigla de la referencia
        "codigo_oficio": codigo_oficio,    # Referencia oficio
        "tipo_accion": tipo_accion,
        "causal_oficio": causal_oficio,
        "fecha_recepcion": fecha_recepcion,
        "fecha_oficio": fecha_oficio,
        "fecha_asignacion": fecha_asignacion,
        "fecha_respuesta": fecha_respuesta,
        "cantidad_investigados": cantidad_investigados,
        "prioridad": prioridad,
        "implicados": detalle,
        "id_empleado": id_empleado,
        "empleado": nombre_empleado,
        "estado": estado,
        "observacion": observacion,
        "archivo_oficio": archivo_oficio,      # documento del oficio (PDF/Word)
        "archivo_respuesta": archivo_respuesta,  # PDF de respuesta adjunto
        "registrado_por": registrado_por,
        "fecha_registro": ahora,
        "historial": [{"estado": estado, "por": registrado_por, "cuando": ahora}],
    })
    _guardar_registros(registros)
    registro_actividad.registrar(
        "REGISTRAR_OFICIO",
        f"referencia={referencia}; codigo={codigo_oficio}; "
        f"responsable={nombre_empleado or '(sin responsable)'}; estado={estado}",
        registrado_por)
    return referencia


def validar_importacion(filas: List[Dict], actor: str = "",
                        actor_rol: str = None) -> List[Dict]:
    """Comprueba TODAS las filas como si se fueran a guardar, sin guardar nada.

    Es la misma comprobación que hace `importar_oficios`, hecha por adelantado
    para que la vista previa pueda mostrar las filas con error ANTES de tocar
    los datos. Devuelve una entrada por oficio que no pasa (ver
    `carga_masiva.error_de_fila`); lista vacía significa que el archivo entero
    se puede importar.
    """
    from carga_masiva import error_de_fila, etiqueta_filas

    registros = _leer_registros()
    codigos = {r.get("codigo_oficio", "").strip().casefold() for r in registros}
    ahora = datetime.now().isoformat(timespec="seconds")
    errores = []
    for fila in filas:
        try:
            nuevo = _preparar_importado(fila, registros, codigos, actor,
                                        ahora, actor_rol)
        except ValueError as error:
            errores.append(error_de_fila(etiqueta_filas(fila),
                                         fila.get("codigo_oficio"), error))
            continue
        # Se anota lo ya validado para que el resto del archivo choque con ello:
        # dos oficios con la misma referencia también son un error.
        registros.append(nuevo)
        codigos.add(nuevo["codigo_oficio"].casefold())
    return errores


@bloqueo.con_bloqueo("oficios")
def importar_oficios(filas: List[Dict], importado_por: str,
                     actor_rol: str = None) -> Dict:
    """Da de alta en bloque los oficios de una carga masiva. **Todo o nada.**

    Cada oficio se valida con las mismas reglas que el alta manual (fechas
    coherentes, estado acorde al responsable, responsable existente, tipo de
    acción del catálogo, implicados bien formados, referencias sin repetir). Si
    UNA sola fila falla no se guarda nada y se devuelven todas las que fallan,
    con su línea del archivo y su motivo: un archivo se importa entero o se
    corrige entero, para no dejar una carga a medias que nadie sabe dónde
    quedó.

    Lo único que no se exige, porque un archivo no puede aportarlo, es el
    documento del oficio y la respuesta en PDF; se adjuntan después desde la
    pestaña Oficios. Un oficio que llegue como "Finalizado" sí tiene que traer
    sus fechas de asignación y de respuesta.

    Se hace en una sola escritura: la carga entera se guarda de golpe.
    """
    if actor_rol is not None and actor_rol not in ROLES_GESTORES:
        raise ValueError(
            "La carga masiva de oficios está reservada a administradores y al "
            "superusuario."
        )
    # Import diferido: quien lee el archivo sabe qué líneas componen cada
    # oficio, y así el aviso señala la línea que hay que corregir.
    from carga_masiva import error_de_fila, etiqueta_filas

    registros = _leer_registros()
    codigos = {r.get("codigo_oficio", "").strip().casefold() for r in registros}

    importados, fallidos = [], []
    ahora = datetime.now().isoformat(timespec="seconds")
    for fila in filas:
        try:
            nuevo = _preparar_importado(fila, registros, codigos,
                                        importado_por, ahora, actor_rol)
        except ValueError as error:
            fallidos.append(error_de_fila(etiqueta_filas(fila),
                                          fila.get("codigo_oficio"), error))
            continue
        registros.append(nuevo)
        codigos.add(nuevo["codigo_oficio"].casefold())
        importados.append(nuevo["referencia"])

    if fallidos:
        # No se guarda NADA: el archivo se corrige y se vuelve a cargar entero.
        registro_actividad.registrar(
            "CARGA_MASIVA_RECHAZADA",
            f"oficios={len(filas)}; con errores={len(fallidos)}", importado_por)
        return {"importados": [], "fallidos": fallidos}

    _guardar_registros(registros)
    registro_actividad.registrar(
        "CARGA_MASIVA", f"importados={len(importados)}", importado_por)
    return {"importados": importados, "fallidos": []}


def _empleado_de(id_empleado: str) -> Dict:
    """Cuenta del responsable indicado en el archivo.

    El archivo nombra al responsable por su NOMBRE DE CUENTA (la columna
    «Usuario responsable» de la exportación), que es lo único que identifica a
    una persona sin ambigüedad. Si no existe, se dice para que se cree antes de
    volver a cargar: la carga no inventa usuarios.
    """
    import autenticacion
    objetivo = (id_empleado or "").strip().casefold()
    for registro in autenticacion.listar_usuarios():
        if registro["usuario"].strip().casefold() == objetivo:
            return registro
    raise ValueError(
        f"el usuario «{id_empleado}» no existe en el sistema. Créelo primero "
        f"en la pestaña Usuarios y vuelva a cargar el archivo."
    )


def _validar_anulacion(anulado: str, motivo: str) -> Tuple[bool, str]:
    """Columnas «Anulado» y «Motivo de anulación» del archivo."""
    valor = " ".join(str(anulado or "").split()).casefold()
    motivo = " ".join(str(motivo or "").split())
    if valor in ("", "no", "false", "0"):
        if motivo:
            raise ValueError(
                "hay un motivo de anulación pero el oficio no está marcado "
                "como anulado.")
        return False, ""
    if valor not in ("si", "sí", "true", "1"):
        raise ValueError(
            f"«{anulado}» no es un valor de «Anulado» válido. Opciones: Sí o No.")
    if len(motivo) < 5:
        raise ValueError("indique el motivo de la anulación.")
    return True, motivo


def _preparar_importado(fila: Dict, registros: List[Dict], codigos: set,
                        importado_por: str, ahora: str,
                        actor_rol: str = None) -> Dict:
    """Valida un oficio de la carga masiva y devuelve el registro a guardar.

    Aplica las reglas del alta manual: lo que la aplicación no dejaría
    registrar a mano tampoco entra por el archivo.
    """
    codigo_oficio = (fila.get("codigo_oficio") or "").strip()
    if not codigo_oficio:
        raise ValueError("falta la Referencia oficio.")
    if codigo_oficio.casefold() in codigos:
        raise ValueError(
            f"el oficio «{codigo_oficio}» ya está registrado. Quítelo del "
            f"archivo: la carga da de alta oficios nuevos, no actualiza los "
            f"existentes.")

    # La Referencia UDC no se toma del archivo: la genera el sistema con la
    # nomenclatura de la institución que remite el oficio.
    institucion = parametros.validar_institucion(fila.get("institucion"))
    tipo_accion = _validar_tipo_accion(fila.get("tipo_accion"))
    referencia = _generar_referencia(registros, institucion)

    fecha_recepcion = (fila.get("fecha_recepcion") or "").strip()
    if not fecha_recepcion:
        raise ValueError("falta la fecha de recepción.")
    _validar_fecha(fecha_recepcion, "Fecha de recepción")

    fecha_oficio = (fila.get("fecha_oficio") or "").strip()
    if not fecha_oficio:
        raise ValueError("falta la fecha de oficio.")
    _validar_fecha(fecha_oficio, "Fecha de oficio")
    if datetime.strptime(fecha_oficio, "%Y-%m-%d") > datetime.strptime(fecha_recepcion, "%Y-%m-%d"):
        raise ValueError(
            "la fecha de oficio no puede ser posterior a la de recepción.")

    fecha_respuesta = _validar_fecha_respuesta(
        fila.get("fecha_respuesta"), fecha_recepcion)
    fecha_asignacion = _validar_fecha_asignacion(
        fila.get("fecha_asignacion"), fecha_recepcion)
    prioridad = _validar_prioridad(fila.get("prioridad"))
    anulado, motivo_anulacion = _validar_anulacion(
        fila.get("anulado"), fila.get("motivo_anulacion"))

    # Cada fila del archivo es una persona investigada, así que el detalle llega
    # con el oficio y es quien manda sobre la cantidad.
    implicados = []
    for numero, datos in enumerate(fila.get("implicados") or [], start=1):
        implicado = validar_implicado(
            datos.get("nombre", ""), datos.get("tipo_identificacion", ""),
            datos.get("identificacion", ""), datos.get("tipo_implicado", ""),
            datos.get("lci", "No") or "No")
        implicado["id"] = numero
        implicados.append(implicado)
    cantidad = _validar_cantidad_investigados(fila.get("cantidad_investigados"))
    _exigir_cantidad_coherente({"implicados": implicados}, cantidad)
    if implicados:
        cantidad = str(len(implicados))

    # Responsable: se identifica por su nombre de cuenta y tiene que existir.
    id_empleado = (fila.get("id_empleado") or "").strip()
    nombre_archivo = (fila.get("empleado") or "").strip()
    nombre_empleado = ""
    if id_empleado:
        cuenta = _empleado_de(id_empleado)
        id_empleado = cuenta["usuario"]
        nombre_empleado = cuenta.get("nombre", "")
        _validar_asignacion(id_empleado, actor_rol)
    elif nombre_archivo:
        raise ValueError(
            f"el oficio indica «{nombre_archivo}» como responsable pero no dice "
            f"su usuario. Complete la columna «Usuario responsable».")

    if fecha_asignacion and not id_empleado:
        raise ValueError(
            "tiene fecha de asignación pero no responsable: indique a quién se "
            "le asignó o deje la fecha en blanco.")

    estado = (fila.get("estado") or "").strip()
    if estado not in ESTADOS:
        raise ValueError(
            f"el estado «{estado or '(vacío)'}» no es válido. "
            f"Opciones: {', '.join(ESTADOS)}.")
    # Las mismas reglas del alta manual entre responsable, respuesta y estado.
    # Aquí no se corrige en silencio: si el archivo dice otra cosa, se avisa.
    coherente = _resolver_estado(nombre_empleado, estado, fecha_respuesta)
    if coherente != estado:
        raise ValueError(
            f"el estado «{estado}» no concuerda con el resto del oficio: "
            f"corresponde «{coherente}».")
    # Finalizar exige el expediente completo. La respuesta en PDF se adjunta
    # después —un archivo no puede traerla—, pero sus fechas sí tienen que
    # estar.
    if estado == "Finalizado":
        faltan = [etiqueta for etiqueta, valor in
                  (("la fecha de asignación", fecha_asignacion),
                   ("la fecha de respuesta", fecha_respuesta)) if not valor]
        if faltan:
            raise ValueError(
                f"para estar «Finalizado» falta {' y '.join(faltan)}.")

    nuevo = {
        "referencia": referencia,
        "institucion": institucion,
        "codigo_oficio": codigo_oficio,
        "tipo_accion": tipo_accion,
        "causal_oficio": (fila.get("causal_oficio") or "").strip(),
        "fecha_recepcion": fecha_recepcion,
        "fecha_oficio": fecha_oficio,
        "fecha_asignacion": fecha_asignacion,
        "fecha_respuesta": fecha_respuesta,
        "cantidad_investigados": cantidad,
        "prioridad": prioridad,
        "implicados": implicados,
        "id_empleado": id_empleado,
        "empleado": nombre_empleado,
        "estado": estado,
        "observacion": (fila.get("observacion") or "").strip(),
        "archivo_oficio": "",
        "archivo_respuesta": "",
        "registrado_por": importado_por,
        "fecha_registro": ahora,
        "origen": "carga masiva",
        "historial": [{"estado": estado, "por": importado_por, "cuando": ahora,
                       "evento": "Importado desde archivo"}],
    }
    if anulado:
        nuevo.update({
            "anulado": True,
            "motivo_anulacion": motivo_anulacion,
            "anulado_por": importado_por,
            "fecha_anulacion": ahora,
        })
        nuevo["historial"].append(
            {"evento": f"Anulado: {motivo_anulacion}", "por": importado_por,
             "cuando": ahora})
    return nuevo


def contar_por_tipo_accion(tipo: str) -> int:
    """Cuántos oficios usan ese tipo de acción (para el catálogo)."""
    objetivo = " ".join(str(tipo or "").split()).casefold()
    return sum(1 for r in _leer_registros()
               if " ".join((r.get("tipo_accion", "") or "").split()).casefold()
               == objetivo)


@bloqueo.con_bloqueo("oficios")
def renombrar_tipo_accion(anterior: str, nuevo: str, actor: str) -> int:
    """Propaga el cambio de nombre de un tipo de acción a los oficios que lo
    usaban, para que ninguno quede apuntando a un valor inexistente."""
    objetivo = " ".join(str(anterior or "").split()).casefold()
    registros = _leer_registros()
    actualizados = 0
    for registro in registros:
        actual = " ".join((registro.get("tipo_accion", "") or "").split())
        if actual.casefold() == objetivo and actual != nuevo:
            registro["tipo_accion"] = nuevo
            actualizados += 1
    if actualizados:
        _guardar_registros(registros)
    return actualizados


def proxima_referencia(institucion: str) -> str:
    """Referencia UDC que se asignaría al próximo oficio de esa institución
    (solo informativa)."""
    return _generar_referencia(_leer_registros(), institucion)


def listar_oficios() -> List[Dict]:
    return _leer_registros()


def esta_anulado(registro: Dict) -> bool:
    return bool(registro.get("anulado"))


def listar_oficios_visibles(actor: str, actor_rol: str,
                            incluir_anulados: bool = False) -> List[Dict]:
    """Oficios que puede VER el usuario en sesión.

    - Superusuario y administrador: todos.
    - Usuario regular: solo los que él registró o los que tiene asignados.

    Los oficios ANULADOS quedan fuera salvo que se pidan expresamente, y solo
    los ve un gestor: son registros retirados de la operación diaria que se
    conservan por trazabilidad.
    """
    registros = _leer_registros()
    if actor_rol in ROLES_GESTORES:
        visibles = registros
    else:
        actor_norm = (actor or "").strip().lower()
        visibles = [
            registro for registro in registros
            if (registro.get("id_empleado", "") or "").strip().lower() == actor_norm
            or (registro.get("registrado_por", "") or "").strip().lower() == actor_norm
        ]
    if incluir_anulados and actor_rol in ROLES_GESTORES:
        return visibles
    return [r for r in visibles if not esta_anulado(r)]


@bloqueo.con_bloqueo("oficios")
def actualizar_oficio(referencia: str, nuevo_estado: str, id_empleado: str,
                     nombre_empleado: str, actualizado_por: str,
                     actor_rol: str = None, fecha_respuesta: str = None,
                     observacion: str = None, fecha_asignacion: str = None,
                     cantidad_investigados=None, tipo_accion: str = None,
                     prioridad: str = None) -> str:
    """Actualiza estado, responsable, fecha de respuesta y/o observación de un
    oficio en una sola operación, respetando las reglas de negocio
    (ver `_resolver_estado`).

    Reservado a GESTORES (administrador / superusuario): pueden reasignar el
    responsable y fijar cualquier estado. Los usuarios regulares no pueden usar
    esta vía (ver `actualizar_estado_asignado`).

    Devuelve el estado final aplicado (puede diferir del solicitado si las
    reglas lo ajustaron, p. ej. al asignar responsable a un "Por asignar").
    """
    if actor_rol is not None and actor_rol not in ROLES_GESTORES:
        raise ValueError(
            "No tiene permisos para reasignar el responsable ni cambiar "
            "libremente el estado del oficio."
        )
    if nuevo_estado not in ESTADOS:
        raise ValueError("Estado no válido.")
    nombre_empleado = (nombre_empleado or "").strip()
    id_empleado = (id_empleado or "").strip()
    registros = _leer_registros()
    for registro in registros:
        if registro["referencia"] == referencia:
            _exigir_no_anulado(registro)
            cambios = []
            # La restricción de asignar a un superusuario solo aplica cuando el
            # responsable CAMBIA: un administrador puede seguir editando la
            # observación de un oficio que ya estaba asignado a un superusuario.
            if id_empleado != (registro.get("id_empleado", "") or "").strip():
                _validar_asignacion(id_empleado, actor_rol)
            # La fecha de respuesta se resuelve primero: si el oficio queda con
            # respuesta, el estado pasa obligatoriamente a "Finalizado".
            if fecha_respuesta is not None:
                nueva_fecha = _validar_fecha_respuesta(
                    fecha_respuesta, registro["fecha_recepcion"])
            else:
                nueva_fecha = registro.get("fecha_respuesta", "")
            estado_final = _resolver_estado(nombre_empleado, nuevo_estado, nueva_fecha)
            # La fecha de asignación se resuelve ANTES de comprobar si se puede
            # finalizar: puede venir en esta misma llamada.
            if fecha_asignacion is not None:
                nueva_asignacion = _validar_fecha_asignacion(
                    fecha_asignacion, registro["fecha_recepcion"])
            else:
                nueva_asignacion = registro.get("fecha_asignacion", "")
            _exigir_datos_para_finalizar(
                estado_final, registro.get("archivo_respuesta", ""),
                nueva_asignacion, nueva_fecha, registro.get("estado", ""))
            if fecha_asignacion is not None:
                if nueva_asignacion != registro.get("fecha_asignacion", ""):
                    registro["fecha_asignacion"] = nueva_asignacion
                    cambios.append(
                        f"Fecha de asignación: {nueva_asignacion or '(sin fecha)'}")
            if cantidad_investigados is not None:
                nueva_cantidad = _validar_cantidad_investigados(cantidad_investigados)
                _exigir_cantidad_coherente(registro, nueva_cantidad)
                if nueva_cantidad != registro.get("cantidad_investigados", ""):
                    registro["cantidad_investigados"] = nueva_cantidad
                    cambios.append(
                        f"Cantidad de investigados: {nueva_cantidad or '(sin dato)'}")
            if tipo_accion is not None:
                nuevo_tipo = _validar_tipo_accion(tipo_accion)
                if nuevo_tipo != registro.get("tipo_accion", ""):
                    registro["tipo_accion"] = nuevo_tipo
                    cambios.append(f"Tipo de acción: {nuevo_tipo}")
            if prioridad is not None:
                nueva_prioridad = _validar_prioridad(prioridad)
                if nueva_prioridad != registro.get("prioridad", ""):
                    registro["prioridad"] = nueva_prioridad
                    cambios.append(
                        f"Prioridad: {nueva_prioridad or '(sin dato)'}")
            if fecha_respuesta is not None:
                if nueva_fecha != registro.get("fecha_respuesta", ""):
                    registro["fecha_respuesta"] = nueva_fecha
                    cambios.append(f"Fecha de respuesta: {nueva_fecha or '(sin fecha)'}")
            if observacion is not None:
                nueva_obs = observacion.strip()
                if nueva_obs != registro.get("observacion", ""):
                    registro["observacion"] = nueva_obs
                    cambios.append("Observación actualizada")
            if nombre_empleado != registro.get("empleado", ""):
                registro["id_empleado"] = id_empleado
                registro["empleado"] = nombre_empleado
                cambios.append(
                    f"Responsable: {nombre_empleado or '(sin responsable)'}"
                )
            if estado_final != registro.get("estado"):
                registro["estado"] = estado_final
                cambios.append(f"Estado: {estado_final}")
            if cambios:
                registro.setdefault("historial", []).append({
                    "evento": " · ".join(cambios),
                    "por": actualizado_por,
                    "cuando": datetime.now().isoformat(timespec="seconds"),
                })
                _guardar_registros(registros)
                registro_actividad.registrar(
                    "ACTUALIZAR_OFICIO",
                    f"referencia={referencia}; " + "; ".join(cambios),
                    actualizado_por)
            return estado_final
    raise ValueError("No se encontró la referencia indicada.")


@bloqueo.con_bloqueo("oficios")
def actualizar_estado_asignado(referencia: str, actor: str, nuevo_estado: str,
                               fecha_respuesta: str = None,
                               observacion: str = None,
                               cantidad_investigados=None,
                               tipo_accion: str = None,
                               prioridad: str = None) -> str:
    """Actualización desde el rol de usuario regular, sobre sus propios oficios.

    Puede cambiar la fecha de respuesta, la observación y alternar el estado
    entre "En proceso" y "Finalizado" (por si finalizó por error y quiere
    reabrirlo). No puede reasignar el responsable ni dejarlo en "Por asignar".

    Si el oficio queda con fecha de respuesta, el estado pasa siempre a
    "Finalizado": para reabrirlo hay que borrar antes esa fecha.
    """
    estados_permitidos = ("En proceso", "Finalizado")
    if nuevo_estado not in estados_permitidos:
        raise ValueError(
            "Como usuario solo puede alternar entre \"En proceso\" y \"Finalizado\"."
        )
    actor_norm = (actor or "").strip().lower()
    registros = _leer_registros()
    for registro in registros:
        if registro["referencia"] == referencia:
            responsable = (registro.get("id_empleado", "") or "").strip().lower()
            if not responsable or responsable != actor_norm:
                raise ValueError("Solo puede modificar oficios asignados a usted.")
            _exigir_no_anulado(registro)
            cambios = []
            # La fecha de respuesta manda sobre el estado.
            if fecha_respuesta is not None:
                nueva_fecha = _validar_fecha_respuesta(
                    fecha_respuesta, registro["fecha_recepcion"])
            else:
                nueva_fecha = registro.get("fecha_respuesta", "")
            estado_final = _resolver_estado(
                registro.get("empleado", ""), nuevo_estado, nueva_fecha)
            # El usuario regular no maneja la fecha de asignación: se comprueba
            # la que ya tenga el oficio (la pone un gestor al asignárselo).
            _exigir_datos_para_finalizar(
                estado_final, registro.get("archivo_respuesta", ""),
                registro.get("fecha_asignacion", ""), nueva_fecha,
                registro.get("estado", ""))
            if cantidad_investigados is not None:
                nueva_cantidad = _validar_cantidad_investigados(cantidad_investigados)
                _exigir_cantidad_coherente(registro, nueva_cantidad)
                if nueva_cantidad != registro.get("cantidad_investigados", ""):
                    registro["cantidad_investigados"] = nueva_cantidad
                    cambios.append(
                        f"Cantidad de investigados: {nueva_cantidad or '(sin dato)'}")
            if tipo_accion is not None:
                nuevo_tipo = _validar_tipo_accion(tipo_accion)
                if nuevo_tipo != registro.get("tipo_accion", ""):
                    registro["tipo_accion"] = nuevo_tipo
                    cambios.append(f"Tipo de acción: {nuevo_tipo}")
            if prioridad is not None:
                nueva_prioridad = _validar_prioridad(prioridad)
                if nueva_prioridad != registro.get("prioridad", ""):
                    registro["prioridad"] = nueva_prioridad
                    cambios.append(
                        f"Prioridad: {nueva_prioridad or '(sin dato)'}")
            if fecha_respuesta is not None:
                if nueva_fecha != registro.get("fecha_respuesta", ""):
                    registro["fecha_respuesta"] = nueva_fecha
                    cambios.append(f"Fecha de respuesta: {nueva_fecha or '(sin fecha)'}")
            if observacion is not None:
                nueva_obs = observacion.strip()
                if nueva_obs != registro.get("observacion", ""):
                    registro["observacion"] = nueva_obs
                    cambios.append("Observación actualizada")
            if registro.get("estado") != estado_final:
                registro["estado"] = estado_final
                cambios.append(f"Estado: {estado_final}")
            if cambios:
                registro.setdefault("historial", []).append({
                    "evento": " · ".join(cambios),
                    "por": actor,
                    "cuando": datetime.now().isoformat(timespec="seconds"),
                })
                _guardar_registros(registros)
                registro_actividad.registrar(
                    "ACTUALIZAR_OFICIO",
                    f"referencia={referencia}; " + "; ".join(cambios), actor)
            return estado_final
    raise ValueError("No se encontró la referencia indicada.")


# --- Mantenimiento: corrección de datos y anulación --------------------------
# Campos que solo se pueden corregir desde el mantenimiento, porque identifican
# al oficio y no forman parte de su trámite diario.
CAMPOS_MANTENIMIENTO = ("codigo_oficio", "causal_oficio",
                        "fecha_oficio", "fecha_recepcion")


@bloqueo.con_bloqueo("oficios")
def corregir_oficio(referencia: str, actor: str, actor_rol: str,
                    **campos) -> List[str]:
    """Corrige los datos de identificación de un oficio mal registrado.

    Reservado a GESTORES (administrador y superusuario). Solo acepta los campos
    de `CAMPOS_MANTENIMIENTO`; el resto del trámite (estado, responsable,
    fechas de asignación y respuesta, observación) se cambia desde el panel
    normal de la pestaña Oficios.

    Se validan las mismas reglas que al registrar: la referencia del oficio no
    puede repetirse y las fechas mantienen su orden. Devuelve la lista de
    cambios aplicados.
    """
    if actor_rol not in ROLES_GESTORES:
        raise ValueError(
            "El mantenimiento de oficios está reservado a administradores y al "
            "superusuario."
        )
    desconocidos = set(campos) - set(CAMPOS_MANTENIMIENTO)
    if desconocidos:
        raise ValueError(
            f"No se pueden corregir estos campos: {', '.join(sorted(desconocidos))}.")

    registros = _leer_registros()
    for registro in registros:
        if registro["referencia"] != referencia:
            continue

        # Los valores finales, mezclando lo que se corrige con lo que ya había.
        nuevo = {campo: (campos[campo] if campo in campos
                         else registro.get(campo, ""))
                 for campo in CAMPOS_MANTENIMIENTO}

        codigo = (nuevo["codigo_oficio"] or "").strip()
        if not codigo:
            raise ValueError("Debe indicar la referencia del oficio.")
        for otro in registros:
            if (otro is not registro and not esta_anulado(otro)
                    and otro.get("codigo_oficio", "").strip().casefold()
                    == codigo.casefold()):
                raise ValueError(
                    f"Ya existe otro oficio con la referencia «{codigo}».")

        recepcion = _validar_fecha((nuevo["fecha_recepcion"] or "").strip(),
                                   "Fecha de recepción")
        oficio = _validar_fecha((nuevo["fecha_oficio"] or "").strip(),
                                "Fecha de oficio")
        if datetime.strptime(oficio, "%Y-%m-%d") > datetime.strptime(recepcion, "%Y-%m-%d"):
            raise ValueError(
                "La fecha de oficio no puede ser posterior a la de recepción.")
        # Cambiar la recepción puede dejar incoherentes las otras dos fechas.
        _validar_fecha_asignacion(registro.get("fecha_asignacion", ""), recepcion)
        _validar_fecha_respuesta(registro.get("fecha_respuesta", ""), recepcion)

        etiquetas = {"codigo_oficio": "Referencia oficio",
                     "causal_oficio": "Causal oficio",
                     "fecha_oficio": "Fecha de oficio",
                     "fecha_recepcion": "Fecha de recepción"}
        cambios = []
        for campo in CAMPOS_MANTENIMIENTO:
            valor = (nuevo[campo] or "").strip()
            if valor != (registro.get(campo, "") or ""):
                cambios.append(f"{etiquetas[campo]}: "
                               f"«{registro.get(campo, '') or '(vacío)'}» → «{valor}»")
                registro[campo] = valor
        if cambios:
            registro.setdefault("historial", []).append({
                "evento": "Corrección: " + " · ".join(cambios),
                "por": actor,
                "cuando": datetime.now().isoformat(timespec="seconds"),
            })
            _guardar_registros(registros)
            registro_actividad.registrar(
                "CORREGIR_OFICIO",
                f"referencia={referencia}; " + "; ".join(cambios), actor)
        return cambios
    raise ValueError("No se encontró la referencia indicada.")


@bloqueo.con_bloqueo("oficios")
def anular_oficio(referencia: str, motivo: str, actor: str,
                  actor_rol: str) -> None:
    """Retira un oficio de la operación sin borrarlo.

    No se elimina a propósito: la Referencia UDC no se reutiliza, así que un
    borrado real dejaría un hueco en la numeración imposible de explicar, y en
    una unidad de cumplimiento un registro que desaparece sin rastro es difícil
    de justificar. El oficio queda marcado, con su motivo y su autor, fuera del
    listado y de las métricas, y se puede reactivar.
    """
    if actor_rol not in ROLES_GESTORES:
        raise ValueError(
            "Anular oficios está reservado a administradores y al superusuario.")
    motivo = " ".join((motivo or "").split())
    if len(motivo) < 5:
        raise ValueError("Indique el motivo de la anulación.")

    registros = _leer_registros()
    for registro in registros:
        if registro["referencia"] == referencia:
            if esta_anulado(registro):
                raise ValueError("Este oficio ya está anulado.")
            ahora = datetime.now().isoformat(timespec="seconds")
            registro["anulado"] = True
            registro["motivo_anulacion"] = motivo
            registro["anulado_por"] = actor
            registro["fecha_anulacion"] = ahora
            registro.setdefault("historial", []).append({
                "evento": f"Anulado: {motivo}", "por": actor, "cuando": ahora})
            _guardar_registros(registros)
            registro_actividad.registrar(
                "ANULAR_OFICIO", f"referencia={referencia}; motivo={motivo}", actor)
            return
    raise ValueError("No se encontró la referencia indicada.")


@bloqueo.con_bloqueo("oficios")
def reactivar_oficio(referencia: str, actor: str, actor_rol: str) -> None:
    """Devuelve a la operación un oficio anulado por error."""
    if actor_rol not in ROLES_GESTORES:
        raise ValueError(
            "Reactivar oficios está reservado a administradores y al superusuario.")
    registros = _leer_registros()
    for registro in registros:
        if registro["referencia"] == referencia:
            if not esta_anulado(registro):
                raise ValueError("Este oficio no está anulado.")
            ahora = datetime.now().isoformat(timespec="seconds")
            registro["anulado"] = False
            registro["motivo_anulacion"] = ""
            registro.setdefault("historial", []).append({
                "evento": "Reactivado", "por": actor, "cuando": ahora})
            _guardar_registros(registros)
            registro_actividad.registrar(
                "REACTIVAR_OFICIO", f"referencia={referencia}", actor)
            return
    raise ValueError("No se encontró la referencia indicada.")


# --- Respuesta en PDF adjunta ------------------------------------------------
def _puede_editar(registro: Dict, actor: str, actor_rol: str) -> bool:
    """Un gestor puede sobre cualquier oficio; un usuario regular solo sobre
    los oficios asignados a él."""
    if actor_rol in ROLES_GESTORES:
        return True
    responsable = (registro.get("id_empleado", "") or "").strip().lower()
    return bool(responsable) and responsable == (actor or "").strip().lower()


def _ruta_adjunto(referencia: str, campo: str, carpeta: Path) -> Optional[Path]:
    for registro in _leer_registros():
        if registro["referencia"] == referencia:
            nombre = registro.get(campo, "")
            if not nombre:
                return None
            ruta = carpeta / nombre
            return ruta if ruta.exists() else None
    return None


def ruta_respuesta(referencia: str) -> Optional[Path]:
    """Devuelve la ruta del PDF de respuesta adjunto, o None si no hay."""
    return _ruta_adjunto(referencia, "archivo_respuesta", DIR_RESPUESTAS)


def ruta_documento(referencia: str) -> Optional[Path]:
    """Devuelve la ruta del documento del oficio (PDF o Word), o None."""
    return _ruta_adjunto(referencia, "archivo_oficio", DIR_DOCUMENTOS)


@bloqueo.con_bloqueo("oficios")
def reemplazar_documento(referencia: str, ruta_origen: str, actor: str,
                         actor_rol: str) -> str:
    """Sustituye el documento del oficio (PDF o Word) por si se cargó el
    archivo equivocado. Mismos permisos que adjuntar la respuesta."""
    registros = _leer_registros()
    for registro in registros:
        if registro["referencia"] == referencia:
            if not _puede_editar(registro, actor, actor_rol):
                raise ValueError(
                    "Solo puede cambiar el documento de oficios asignados a usted.")
            anterior = registro.get("archivo_oficio", "")
            nombre = _guardar_documento(
                referencia, ruta_origen, DIR_DOCUMENTOS, EXTENSIONES_DOCUMENTO,
                "El documento del oficio")
            # Si la extensión cambió (.pdf -> .docx), retirar el archivo previo.
            if anterior and anterior != nombre:
                try:
                    permisos.hacer_escribible(DIR_DOCUMENTOS / anterior)
                    (DIR_DOCUMENTOS / anterior).unlink(missing_ok=True)
                except OSError:
                    pass
            registro["archivo_oficio"] = nombre
            registro.setdefault("historial", []).append({
                "evento": "Documento del oficio actualizado",
                "por": actor,
                "cuando": datetime.now().isoformat(timespec="seconds"),
            })
            _guardar_registros(registros)
            registro_actividad.registrar(
                "REEMPLAZAR_DOCUMENTO",
                f"referencia={referencia}; archivo={nombre}", actor)
            return nombre
    raise ValueError("No se encontró la referencia indicada.")


@bloqueo.con_bloqueo("oficios")
def adjuntar_respuesta(referencia: str, ruta_origen: str, actor: str,
                       actor_rol: str) -> str:
    """Copia un PDF de respuesta a datos/respuestas/ y lo asocia al oficio.

    El archivo se guarda como '<referencia>.pdf'. Si ya había uno, se
    reemplaza. Devuelve el nombre del archivo guardado.
    """
    registros = _leer_registros()
    for registro in registros:
        if registro["referencia"] == referencia:
            if not _puede_editar(registro, actor, actor_rol):
                raise ValueError("Solo puede adjuntar respuestas a oficios asignados a usted.")
            nombre = _guardar_documento(
                referencia, ruta_origen, DIR_RESPUESTAS, (".pdf",), "La respuesta")
            registro["archivo_respuesta"] = nombre
            registro.setdefault("historial", []).append({
                "evento": "Respuesta en PDF adjuntada",
                "por": actor,
                "cuando": datetime.now().isoformat(timespec="seconds"),
            })
            _guardar_registros(registros)
            registro_actividad.registrar(
                "ADJUNTAR_RESPUESTA",
                f"referencia={referencia}; archivo={nombre}", actor)
            return nombre
    raise ValueError("No se encontró la referencia indicada.")


@bloqueo.con_bloqueo("oficios")
def eliminar_respuesta(referencia: str, actor: str, actor_rol: str) -> None:
    """Elimina el PDF de respuesta adjunto (por si se cargó el archivo
    equivocado). Mismos permisos que adjuntar: un usuario regular solo sobre
    los oficios asignados a él."""
    registros = _leer_registros()
    for registro in registros:
        if registro["referencia"] == referencia:
            if not _puede_editar(registro, actor, actor_rol):
                raise ValueError("Solo puede eliminar respuestas de oficios asignados a usted.")
            nombre = registro.get("archivo_respuesta", "")
            if not nombre:
                raise ValueError("Este oficio no tiene una respuesta en PDF adjunta.")
            # Un oficio finalizado no puede quedarse sin su respuesta: primero
            # hay que reabrirlo (borrando la fecha de respuesta).
            if registro.get("estado") == "Finalizado":
                raise ValueError(
                    "No se puede quitar la respuesta de un oficio finalizado. "
                    "Reabra antes el oficio borrando su fecha de respuesta."
                )
            ruta = DIR_RESPUESTAS / nombre
            try:
                permisos.hacer_escribible(ruta)
                ruta.unlink(missing_ok=True)
            except OSError as error:
                raise ValueError(f"No se pudo eliminar el PDF: {error}")
            registro["archivo_respuesta"] = ""
            registro.setdefault("historial", []).append({
                "evento": "Respuesta en PDF eliminada",
                "por": actor,
                "cuando": datetime.now().isoformat(timespec="seconds"),
            })
            _guardar_registros(registros)
            registro_actividad.registrar(
                "ELIMINAR_RESPUESTA",
                f"referencia={referencia}; archivo={nombre}", actor)
            return
    raise ValueError("No se encontró la referencia indicada.")


# --- Búsqueda / filtros ------------------------------------------------------
# Campos de texto por los que se puede buscar: clave interna -> etiqueta.
CAMPOS_BUSQUEDA = {
    "referencia": "Referencia UDC",
    "institucion": "Institución del Estado",
    "codigo_oficio": "Referencia oficio",
    "tipo_accion": "Tipo de acción",
    "causal_oficio": "Causal oficio",
}

# Tipos de fecha por los que se puede filtrar (y exportar).
CAMPOS_FECHA = {
    "fecha_oficio": "Fecha de oficio",
    "fecha_recepcion": "Fecha de recepción",
    "fecha_asignacion": "Fecha de asignación",
    "fecha_respuesta": "Fecha de respuesta",
}

# Columnas del CSV de exportación: clave interna -> encabezado.
# Columnas de la exportación: TODO lo que se captura del oficio, en el orden en
# que se pide en los formularios.
COLUMNAS_EXPORTACION = {
    "referencia": "Referencia UDC",
    "institucion": "Institución del Estado",
    "codigo_oficio": "Referencia oficio",
    "tipo_accion": "Tipo de acción",
    "causal_oficio": "Causal oficio",
    "fecha_oficio": "Fecha de oficio",
    "fecha_recepcion": "Fecha de recepción",
    "fecha_asignacion": "Fecha de asignación",
    "fecha_respuesta": "Fecha de respuesta",
    "cantidad_investigados": "Cantidad de investigados",
    "prioridad": "Prioridad",
    "id_empleado": "Usuario responsable",
    "empleado": "Responsable",
    "estado": "Estado",
    "archivo_oficio": "Documento del oficio",
    "archivo_respuesta": "Respuesta en PDF",
    "observacion": "Observación",
    "registrado_por": "Registrado por",
    "fecha_registro": "Fecha de registro",
    "origen": "Origen",
    "anulado": "Anulado",
    "motivo_anulacion": "Motivo de anulación",
}

# Columnas del implicado que se añaden a la derecha de las del oficio.
COLUMNAS_IMPLICADO = {
    "nombre": "Implicado",
    "tipo_identificacion": "Tipo de identificación",
    "identificacion": "Identificación",
    "tipo_implicado": "Tipo de implicado",
    "lci": "LCI",
}


# Separador del CSV exportado.
#
# Se usa la barra vertical en lugar de ';' o ',' porque los campos de texto
# libre del oficio (observación, causal) pueden contener comas y puntos y coma,
# y algunos programas los interpretan como separador aunque el valor venga
# entrecomillado, partiendo la fila en columnas equivocadas. La barra vertical
# no aparece en la práctica en el texto de un oficio, así que la importación es
# inequívoca. Es un detalle interno del formato: no se menciona en la interfaz.
SEPARADOR_CSV = "|"


def _fila_exportacion(registro: Dict) -> List[str]:
    """Valores de un oficio en el orden de COLUMNAS_EXPORTACION, con los saltos
    de línea de la observación colapsados en espacios."""
    valores = []
    for clave in COLUMNAS_EXPORTACION:
        valor = registro.get(clave, "")
        if isinstance(valor, bool):        # "anulado" se lee mejor así
            valor = "Sí" if valor else "No"
        valores.append(" ".join(str(valor or "").split()))
    return valores


def _encabezados_exportacion() -> List[str]:
    return list(COLUMNAS_EXPORTACION.values()) + list(COLUMNAS_IMPLICADO.values())


def filas_exportacion(registros: List[Dict]) -> List[List[str]]:
    """Todas las filas de la exportación: **una por implicado**.

    Cada oficio se repite tantas veces como personas investiga, igual que en la
    matriz de la unidad, con los datos del oficio a la izquierda y los del
    implicado a la derecha. Un oficio sin implicados anotados ocupa una sola
    fila, con esas últimas columnas vacías: no se pierde del reporte.
    """
    filas = []
    for registro in registros:
        datos = _fila_exportacion(registro)
        implicados = registro.get("implicados") or []
        if not implicados:
            filas.append(datos + [""] * len(COLUMNAS_IMPLICADO))
            continue
        for implicado in implicados:
            filas.append(datos + [
                " ".join(str(implicado.get(clave, "") or "").split())
                for clave in COLUMNAS_IMPLICADO])
    return filas


def exportar_csv(registros: List[Dict], ruta_destino: str) -> None:
    """Escribe los oficios en un CSV.

    UTF-8 con BOM para que Excel respete las tildes, y `SEPARADOR_CSV` como
    delimitador.
    """
    destino = Path(ruta_destino)
    with destino.open("w", newline="", encoding="utf-8-sig") as archivo:
        escritor = csv.writer(archivo, delimiter=SEPARADOR_CSV)
        escritor.writerow(_encabezados_exportacion())
        for fila in filas_exportacion(registros):
            escritor.writerow(fila)


def hay_soporte_xlsx() -> bool:
    """¿Está disponible openpyxl para exportar a Excel?"""
    try:
        import openpyxl                      # noqa: F401
        return True
    except ImportError:
        return False


def exportar_xlsx(registros: List[Dict], ruta_destino: str) -> None:
    """Escribe los oficios en un libro de Excel (.xlsx)."""
    escribir_xlsx(_encabezados_exportacion(), filas_exportacion(registros),
                  ruta_destino)


def escribir_xlsx(encabezados: List[str], filas: List[List[str]],
                  ruta_destino: str, hoja_titulo: str = "Oficios") -> None:
    """Escribe una hoja de cálculo con la cabecera destacada y anchos al dato.

    Lo usan tanto la exportación como el archivo de ejemplo de la carga, para
    que los dos se vean igual y ninguno tenga su propia idea del formato.

    Requiere openpyxl, que es opcional: si falta, se avisa para que se use el
    CSV, que no necesita ninguna librería externa.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ValueError(
            "Para exportar a Excel hace falta la librería openpyxl:\n\n"
            "    pip install openpyxl\n\n"
            "Mientras tanto puede exportar en formato CSV."
        )

    libro = Workbook()
    hoja = libro.active
    hoja.title = hoja_titulo

    hoja.append(list(encabezados))
    # Cabecera con los colores corporativos, para que se distinga de los datos.
    relleno = PatternFill("solid", fgColor="152342")
    for columna in range(1, len(encabezados) + 1):
        celda = hoja.cell(row=1, column=columna)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = relleno
        celda.alignment = Alignment(vertical="center")

    for fila in filas:
        hoja.append(fila)

    # Ancho de columna aproximado al contenido, acotado para que la observación
    # no desborde la pantalla.
    for indice, encabezado in enumerate(encabezados, start=1):
        ancho = max([len(encabezado)]
                    + [len(str(fila[indice - 1])) for fila in filas])
        hoja.column_dimensions[get_column_letter(indice)].width = min(ancho + 2, 45)

    hoja.freeze_panes = "A2"          # la cabecera queda fija al desplazarse
    hoja.auto_filter.ref = hoja.dimensions
    libro.save(ruta_destino)


# Formatos que ofrece el desplegable de exportación: etiqueta -> extensión.
FORMATOS_EXPORTACION = {
    "Excel (.xlsx)": ".xlsx",
    "CSV (.csv)": ".csv",
}


def exportar_oficios(registros: List[Dict], ruta_destino: str,
                     formato: str = ".xlsx", exportado_por: str = "",
                     detalle: str = "") -> int:
    """Exporta los oficios al formato indicado y devuelve cuántos escribió.

    El archivo lleva una fila por implicado (ver `filas_exportacion`), pero lo
    que se cuenta e informa son los OFICIOS: es lo que el usuario pidió
    exportar.
    """
    if not registros:
        raise ValueError("No hay oficios que exportar con ese criterio.")
    formato = (formato or "").lower()
    if formato not in FORMATOS_EXPORTACION.values():
        raise ValueError("Formato de exportación no válido.")
    destino = Path(ruta_destino)
    try:
        if formato == ".xlsx":
            exportar_xlsx(registros, str(destino))
        else:
            exportar_csv(registros, str(destino))
    except OSError as error:
        raise ValueError(f"No se pudo escribir el archivo: {error}")
    registro_actividad.registrar(
        "EXPORTAR_OFICIOS",
        f"cantidad={len(registros)}; formato={formato}; archivo={destino.name}"
        + (f"; {detalle}" if detalle else ""),
        exportado_por)
    return len(registros)


def filtrar_oficios(registros: List[Dict], campo_texto: str = "", texto: str = "",
                    campo_fecha: str = "", desde: str = "", hasta: str = "",
                    institucion: str = "", tipo_accion: str = "",
                    causal: str = "", estado: str = "", prioridad: str = "",
                    id_empleado: str = "",
                    solo_sin_responsable: bool = False) -> List[Dict]:
    """Filtra una lista de oficios. Todos los filtros se acumulan (Y lógico).

    - `campo_texto` + `texto`: coincidencia parcial, sin distinguir
      mayúsculas/minúsculas, sobre uno de los campos de `CAMPOS_BUSQUEDA`.
    - `campo_fecha` + `desde`/`hasta`: rango sobre UN SOLO tipo de fecha
      (`CAMPOS_FECHA`). Ambos extremos se comparan contra el mismo campo, por
      lo que no es posible mezclar tipos de fecha. Si solo se indica `desde`,
      se busca esa **fecha única**; si solo se indica `hasta`, todo lo anterior
      o igual a esa fecha.
    - `institucion`, `tipo_accion`, `causal`, `estado` y `prioridad`:
      coincidencia **exacta**, porque se eligen de un desplegable con los
      valores que existen.
    - `id_empleado`: oficios de ese responsable. `solo_sin_responsable` los
      deja en los que aún no tienen a nadie a cargo; las dos cosas no se
      combinan (manda `solo_sin_responsable`).

    Los oficios sin la fecha indicada (por ejemplo sin fecha de respuesta) no
    aparecen cuando se filtra por ese campo.
    """
    resultado = registros

    texto = (texto or "").strip().casefold()
    if texto:
        if campo_texto not in CAMPOS_BUSQUEDA:
            raise ValueError("Campo de búsqueda no válido.")
        resultado = [r for r in resultado
                     if texto in (r.get(campo_texto, "") or "").casefold()]

    desde = (desde or "").strip()
    hasta = (hasta or "").strip()
    if desde or hasta:
        if campo_fecha not in CAMPOS_FECHA:
            raise ValueError("Tipo de fecha no válido.")
        etiqueta = CAMPOS_FECHA[campo_fecha]
        if desde:
            _validar_fecha(desde, f"{etiqueta} (desde)")
        if hasta:
            _validar_fecha(hasta, f"{etiqueta} (hasta)")
        if desde and hasta and desde > hasta:
            raise ValueError(
                "La fecha inicial no puede ser posterior a la fecha final."
            )
        # Fecha única: solo se indicó el extremo inicial.
        if desde and not hasta:
            hasta = desde
        filtrados = []
        for registro in resultado:
            valor = (registro.get(campo_fecha, "") or "").strip()
            if not valor:
                continue          # sin esa fecha -> no participa del filtro
            if desde and valor < desde:
                continue
            if hasta and valor > hasta:
                continue
            filtrados.append(registro)
        resultado = filtrados

    institucion = (institucion or "").strip()
    if institucion:
        institucion = parametros.validar_institucion(institucion)
        resultado = [r for r in resultado
                     if (r.get("institucion", "") or "") == institucion]

    tipo_accion = (tipo_accion or "").strip()
    if tipo_accion:
        resultado = [r for r in resultado
                     if (r.get("tipo_accion", "") or "") == tipo_accion]

    causal = (causal or "").strip()
    if causal:
        resultado = [r for r in resultado
                     if (r.get("causal_oficio", "") or "") == causal]

    estado = (estado or "").strip()
    if estado:
        if estado not in ESTADOS:
            raise ValueError(f"El estado «{estado}» no es válido.")
        resultado = [r for r in resultado if r.get("estado", "") == estado]

    prioridad = (prioridad or "").strip()
    if prioridad:
        prioridad = _validar_prioridad(prioridad)
        resultado = [r for r in resultado
                     if (r.get("prioridad", "") or "") == prioridad]

    if solo_sin_responsable:
        resultado = [r for r in resultado if not (r.get("id_empleado") or "").strip()]
    elif (id_empleado or "").strip():
        objetivo = id_empleado.strip().casefold()
        resultado = [r for r in resultado
                     if (r.get("id_empleado", "") or "").casefold() == objetivo]

    return resultado


def causales_registradas(registros: List[Dict]) -> List[str]:
    """Causales distintas presentes en esos oficios, en orden alfabético.

    El causal es texto libre, así que el desplegable del filtro se arma con lo
    que realmente hay registrado en vez de con un catálogo fijo.
    """
    return sorted({(r.get("causal_oficio", "") or "").strip()
                   for r in registros
                   if (r.get("causal_oficio", "") or "").strip()},
                  key=str.casefold)


# --- Implicados (personas investigadas en un oficio) -------------------------
# Un oficio puede pedir información sobre varias personas. Cada una se guarda
# dentro del propio oficio, en la lista "implicados", porque no tienen vida
# fuera de él: son el detalle de ese requerimiento.
#
# Cada implicado lleva un `id` propio (correlativo dentro del oficio) en vez de
# identificarse por su posición: así, si alguien elimina uno mientras otra
# persona edita, no se modifica al que no era.
# Separadores que se admiten al teclear una identificación y que no forman
# parte de ella: "1400.349.096" y "1400349096" son el mismo documento.
_SEPARADORES_IDENTIFICACION = str.maketrans("", "", " .-/")

# Cuántos dígitos tiene cada documento numérico.
DIGITOS_IDENTIFICACION = {"Cédula": 10, "RUC": 13}


def validar_identificacion(tipo_identificacion: str, identificacion: str) -> str:
    """Comprueba la identificación según su tipo y la devuelve normalizada.

    - **Cédula**: 10 dígitos.
    - **RUC**: 13 dígitos.
    - **Pasaporte**: letras y números, sin más restricciones (cada país usa su
      propio formato).

    Los puntos, guiones y espacios se retiran antes de comprobar: son formas de
    escribir el mismo documento. Lo que se guarda es el valor ya limpio.
    """
    identificacion = "".join(str(identificacion or "").split())
    if not identificacion:
        return ""
    limpia = identificacion.translate(_SEPARADORES_IDENTIFICACION)

    digitos = DIGITOS_IDENTIFICACION.get(tipo_identificacion)
    if digitos:
        if not limpia.isdigit():
            raise ValueError(
                f"{tipo_identificacion}: «{identificacion}» debe contener solo "
                f"números."
            )
        if len(limpia) != digitos:
            raise ValueError(
                f"{tipo_identificacion}: debe tener {digitos} dígitos "
                f"(se ingresaron {len(limpia)})."
            )
    elif tipo_identificacion == "Pasaporte":
        if not limpia.isalnum():
            raise ValueError(
                f"Pasaporte: «{identificacion}» solo admite letras y números."
            )
    return limpia


def validar_implicado(nombre: str, tipo_identificacion: str = "",
                      identificacion: str = "", tipo_implicado: str = "",
                      lci: str = "No") -> Dict:
    """Comprueba y normaliza los datos de un implicado."""
    nombre = " ".join(str(nombre or "").split())
    if len(nombre) < 3:
        raise ValueError("Debe ingresar el nombre o razón social del implicado.")

    tipo_identificacion = " ".join(str(tipo_identificacion or "").split())
    if tipo_identificacion and tipo_identificacion not in TIPOS_IDENTIFICACION:
        raise ValueError(
            f"El tipo de identificación «{tipo_identificacion}» no es válido. "
            f"Opciones: {', '.join(TIPOS_IDENTIFICACION)}."
        )
    identificacion = " ".join(str(identificacion or "").split())
    # La identificación es opcional: hay oficios sobre personas de las que la
    # institución no aporta documento (de ahí el tipo "Sin identificación").
    if identificacion and not tipo_identificacion:
        raise ValueError(
            "Indique el tipo de identificación (cédula, pasaporte o RUC)."
        )
    identificacion = validar_identificacion(tipo_identificacion, identificacion)

    tipo_implicado = " ".join(str(tipo_implicado or "").split())
    if tipo_implicado not in TIPOS_IMPLICADO:
        raise ValueError(
            f"Debe indicar el tipo de implicado. "
            f"Opciones: {', '.join(TIPOS_IMPLICADO)}."
        )
    lci = " ".join(str(lci or "").split()) or "No"
    if lci not in VALORES_LCI:
        raise ValueError("El campo LCI solo admite «Sí» o «No».")

    return {
        "nombre": nombre,
        "tipo_identificacion": tipo_identificacion,
        "identificacion": identificacion,
        "tipo_implicado": tipo_implicado,
        "lci": lci,
    }


def _sincronizar_investigados(registro: Dict) -> None:
    """La cantidad de investigados pasa a contarla el detalle.

    Mientras el oficio no tenga implicados anotados, `cantidad_investigados` es
    lo que alguien escribió a mano (o lo que dedujo la carga masiva). En cuanto
    hay detalle, manda el detalle: no tendría sentido decir «3 investigados» y
    listar cuatro personas.
    """
    implicados = registro.get("implicados") or []
    if implicados:
        registro["cantidad_investigados"] = str(len(implicados))


def _oficio_editable(registros: List[Dict], referencia: str, actor: str,
                     actor_rol: str) -> Dict:
    """Devuelve el oficio indicado, comprobando permisos y que no esté anulado."""
    for registro in registros:
        if registro["referencia"] == referencia:
            if not _puede_editar(registro, actor, actor_rol):
                raise ValueError(
                    "Solo puede modificar los implicados de los oficios "
                    "asignados a usted."
                )
            _exigir_no_anulado(registro)
            return registro
    raise ValueError("No se encontró la referencia indicada.")


def listar_implicados(referencia: str) -> List[Dict]:
    """Implicados anotados en un oficio, en el orden en que se registraron."""
    for registro in _leer_registros():
        if registro["referencia"] == referencia:
            return list(registro.get("implicados") or [])
    return []


@bloqueo.con_bloqueo("oficios")
def agregar_implicado(referencia: str, actor: str, actor_rol: str,
                      nombre: str = "", tipo_identificacion: str = "",
                      identificacion: str = "", tipo_implicado: str = "",
                      lci: str = "No") -> Dict:
    """Añade una persona investigada al oficio. Devuelve el implicado guardado."""
    registros = _leer_registros()
    registro = _oficio_editable(registros, referencia, actor, actor_rol)
    implicado = validar_implicado(nombre, tipo_identificacion, identificacion,
                                  tipo_implicado, lci)
    implicados = registro.setdefault("implicados", [])
    implicado["id"] = max((int(i.get("id", 0)) for i in implicados), default=0) + 1
    implicados.append(implicado)
    _sincronizar_investigados(registro)
    registro.setdefault("historial", []).append({
        "evento": f"Implicado añadido: {implicado['nombre']}",
        "por": actor,
        "cuando": datetime.now().isoformat(timespec="seconds"),
    })
    _guardar_registros(registros)
    registro_actividad.registrar(
        "AGREGAR_IMPLICADO",
        f"referencia={referencia}; implicado={implicado['nombre']}", actor)
    return implicado


@bloqueo.con_bloqueo("oficios")
def actualizar_implicado(referencia: str, id_implicado: int, actor: str,
                         actor_rol: str, nombre: str = "",
                         tipo_identificacion: str = "", identificacion: str = "",
                         tipo_implicado: str = "", lci: str = "No") -> Dict:
    """Corrige los datos de un implicado ya anotado."""
    registros = _leer_registros()
    registro = _oficio_editable(registros, referencia, actor, actor_rol)
    nuevos = validar_implicado(nombre, tipo_identificacion, identificacion,
                               tipo_implicado, lci)
    for implicado in registro.get("implicados") or []:
        if int(implicado.get("id", 0)) == int(id_implicado):
            anterior = implicado.get("nombre", "")
            implicado.update(nuevos)
            registro.setdefault("historial", []).append({
                "evento": f"Implicado modificado: {anterior}",
                "por": actor,
                "cuando": datetime.now().isoformat(timespec="seconds"),
            })
            _guardar_registros(registros)
            registro_actividad.registrar(
                "ACTUALIZAR_IMPLICADO",
                f"referencia={referencia}; implicado={nuevos['nombre']}", actor)
            return implicado
    raise ValueError("Ese implicado ya no existe en el oficio.")


@bloqueo.con_bloqueo("oficios")
def eliminar_implicado(referencia: str, id_implicado: int, actor: str,
                       actor_rol: str) -> None:
    """Quita a una persona de la lista de investigados del oficio."""
    registros = _leer_registros()
    registro = _oficio_editable(registros, referencia, actor, actor_rol)
    implicados = registro.get("implicados") or []
    quedan = [i for i in implicados if int(i.get("id", 0)) != int(id_implicado)]
    if len(quedan) == len(implicados):
        raise ValueError("Ese implicado ya no existe en el oficio.")
    eliminado = next(i for i in implicados
                     if int(i.get("id", 0)) == int(id_implicado))
    registro["implicados"] = quedan
    _sincronizar_investigados(registro)
    registro.setdefault("historial", []).append({
        "evento": f"Implicado eliminado: {eliminado.get('nombre', '')}",
        "por": actor,
        "cuando": datetime.now().isoformat(timespec="seconds"),
    })
    _guardar_registros(registros)
    registro_actividad.registrar(
        "ELIMINAR_IMPLICADO",
        f"referencia={referencia}; implicado={eliminado.get('nombre', '')}",
        actor)

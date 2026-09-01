"""
Genera el ARCHIVO DE EJEMPLO de la carga masiva.

Crea `Ejemplo de carga masiva.xlsx`: un archivo pequeño, con el FORMATO
ESTABLECIDO que exige la aplicación (cabecera en la fila 1, de la columna A a
la Z, y los datos desde la fila 2) y unos pocos oficios que muestran los casos
habituales. Sirve como plantilla: se borra el contenido de ejemplo y se
escriben los oficios reales debajo de la cabecera.

    python datos_de_prueba/generar_ejemplo_carga.py

Se diferencia de `generar_datos_prueba.py` en el propósito: aquel produce un
volumen grande para ver el tablero con contenido; este es el modelo del formato.
"""
import sys
from datetime import date, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from carga_masiva import (CABECERA_MATRIZ, FILA_CABECERA,   # noqa: E402
                          PRIMERA_FILA_DATOS)

SALIDA = Path(__file__).resolve().parent / "Ejemplo de carga masiva.xlsx"

# La primera columna solo admite la sigla de la institución.
SB = "SB"
FGE = "FGE"

# Fechas relativas a hoy: el ejemplo no envejece y ninguna queda en el futuro,
# que es algo que la aplicación no admite.
HOY = date.today()


def _dia(dias_atras):
    return HOY - timedelta(days=dias_atras)


# Cada entrada es una FILA del archivo. Los oficios con varias personas
# investigadas ocupan varias filas y repiten la Referencia oficio: así los
# agrupa la aplicación, y de ahí sale la cantidad de investigados.
#
# Los casos que se muestran:
#   1) Oficio de la Superintendencia (SB), finalizado, con una sola persona.
#   2) Oficio de la Fiscalía (FGE), en proceso, con TRES personas investigadas.
#   3) Oficio sin responsable: entra como "Por asignar".
#   4) Oficio de una empresa, identificada con RUC.
#   5) Oficio con pasaporte y sin fecha de respuesta.
FILAS = [
    {
        "Institución del Estado": SB,
        "Fecha Asignación": _dia(40), "Usuario": "C. Roman", "Prioridad": "Alta",
        "Fecha Emisión": _dia(41), "Fecha Envío": _dia(30), "Estado": "Finalizado",
        "Fecha Circular": _dia(45),
        "Apellidos, Nombres - Razón Social": "ORDOÑEZ VILLAGOMEZ DAVID MIGUEL",
        "TiPASo Id CED; PAS; RUCUC": "CED",
        "Identificación Ced; Pas; RUC": "1400349096",
        "Referencia - Oficio FGE; Juzgado": "SB-2026-0101-OF",
        "Delito": "LAVADO DE ACTIVOS", "Tipo de Accion": "CERTIFICACIÓN",
        "Observación": "Atendido dentro del plazo",
        "Tipo de Implicado": "CLIENTE", "LCI - SI o NO": "SI",
    },
    {
        "Institución del Estado": FGE,
        "Fecha Asignación": _dia(20), "Usuario": "J. Portero", "Prioridad": "Media",
        "Fecha Emisión": _dia(21), "Estado": "En proceso", "Fecha Circular": _dia(25),
        "Apellidos, Nombres - Razón Social": "ACOSTA JEREZ DIANA CAROLINA",
        "TiPASo Id CED; PAS; RUCUC": "CED",
        "Identificación Ced; Pas; RUC": "0923847561",
        "Referencia - Oficio FGE; Juzgado": "FPP-FED4-2026-000123-O",
        "Delito": "COHECHO", "Tipo de Accion": "RETENCIÓN",
        "Tipo de Implicado": "CLIENTE", "LCI - SI o NO": "NO",
    },
    {   # misma Referencia oficio: segunda persona del oficio anterior
        "Institución del Estado": FGE,
        "Fecha Asignación": _dia(20), "Usuario": "J. Portero", "Prioridad": "Media",
        "Fecha Emisión": _dia(21), "Estado": "En proceso", "Fecha Circular": _dia(25),
        "Apellidos, Nombres - Razón Social": "MENDOZA SALAS LUIS ALBERTO",
        "TiPASo Id CED; PAS; RUCUC": "CED",
        "Identificación Ced; Pas; RUC": "1712345678",
        "Referencia - Oficio FGE; Juzgado": "FPP-FED4-2026-000123-O",
        "Delito": "COHECHO", "Tipo de Accion": "RETENCIÓN",
        "Tipo de Implicado": "EX CLIENTE", "LCI - SI o NO": "NO",
    },
    {   # tercera persona del mismo oficio
        "Institución del Estado": FGE,
        "Fecha Asignación": _dia(20), "Usuario": "J. Portero", "Prioridad": "Media",
        "Fecha Emisión": _dia(21), "Estado": "En proceso", "Fecha Circular": _dia(25),
        "Apellidos, Nombres - Razón Social": "PARRA NUÑEZ SOFIA ELENA",
        "TiPASo Id CED; PAS; RUCUC": "CED",
        "Identificación Ced; Pas; RUC": "0102938475",
        "Referencia - Oficio FGE; Juzgado": "FPP-FED4-2026-000123-O",
        "Delito": "COHECHO", "Tipo de Accion": "RETENCIÓN",
        "Tipo de Implicado": "NO CLIENTE", "LCI - SI o NO": "SI",
    },
    {   # sin responsable: la aplicación lo deja en "Por asignar"
        "Institución del Estado": SB,
        "Fecha Emisión": _dia(5), "Prioridad": "Baja", "Estado": "Por asignar",
        "Fecha Circular": _dia(8),
        "Apellidos, Nombres - Razón Social": "CEVALLOS MORA JORGE ANDRÉS",
        "TiPASo Id CED; PAS; RUCUC": "CED",
        "Identificación Ced; Pas; RUC": "1309876543",
        "Referencia - Oficio FGE; Juzgado": "SB-2026-0118-OF",
        "Delito": "DEFRAUDACIÓN TRIBUTARIA", "Tipo de Accion": "INFORMACIÓN",
        "Observación": "Pendiente de asignar",
        "Tipo de Implicado": "SIN IDENTIFICACION", "LCI - SI o NO": "NO",
    },
    {   # empresa: se identifica con RUC (13 dígitos)
        "Institución del Estado": SB,
        "Fecha Asignación": _dia(12), "Usuario": "L. Jarrin", "Prioridad": "Alta",
        "Fecha Emisión": _dia(12), "Fecha Envío": _dia(3), "Estado": "Finalizado",
        "Fecha Circular": _dia(15),
        "Apellidos, Nombres - Razón Social": "COMERCIAL LOS ANDES S.A.",
        "TiPASo Id CED; PAS; RUCUC": "RUC",
        "Identificación Ced; Pas; RUC": "1791234567001",
        "Referencia - Oficio FGE; Juzgado": "SB-2026-0125-OF",
        "Delito": "ENRIQUECIMIENTO ILÍCITO", "Tipo de Accion": "BLOQUEO Y RETENCIÓN",
        "Observación": "Se remitió por correo",
        "Tipo de Implicado": "CLIENTE", "LCI - SI o NO": "SI",
    },
    {   # pasaporte: letras y números
        "Institución del Estado": FGE,
        "Fecha Asignación": _dia(9), "Usuario": "D. Franco", "Prioridad": "Media",
        "Fecha Emisión": _dia(10), "Estado": "En proceso", "Fecha Circular": _dia(14),
        "Apellidos, Nombres - Razón Social": "QUISPE ANDRADE PEDRO JOSÉ",
        "TiPASo Id CED; PAS; RUCUC": "PAS",
        "Identificación Ced; Pas; RUC": "AB123456",
        "Referencia - Oficio FGE; Juzgado": "FPP-FED4-2026-000188-O",
        "Delito": "TRAFICO ILÍCITO DE SUSTANCIAS CATALOGADAS SUJETAS A FISCALIZACIÓN",
        "Tipo de Accion": "LEVANTAMIENTO DE MEDIDAS",
        "Tipo de Implicado": "NO CLIENTE", "LCI - SI o NO": "NO",
    },
]

# Columnas que la aplicación no guarda pero que forman parte del formato: se
# rellenan con un valor de muestra para que el archivo se vea como el real.
RELLENO = {
    "Mes": lambda fila: ["ENE", "FEB", "MAR", "ABR", "MAY", "JUN", "JUL", "AGO",
                         "SEP", "OCT", "NOV", "DIC"][fila["Fecha Emisión"].month - 1],
    "Referencia": lambda fila: "",
    "Medio Repuesta": lambda fila: "Electrónico",
    "Días": lambda fila: (fila["Fecha Envío"] - fila["Fecha Asignación"]).days
                         if fila.get("Fecha Envío") and fila.get("Fecha Asignación") else "",
    "Canal Recepc": lambda fila: "Proveedor",
    "Número Expediente Fiscal": lambda fila: "-",
    "Referencia - Circular Superintendencia Bancos": lambda fila: "-",
    "Fecha - Solicitud": lambda fila: fila["Fecha Emisión"] + timedelta(days=1),
    "Ref Solic- No. LCI-202X-000": lambda fila: f"LCI-{HOY.year}-001",
}


def escribir(filas):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Matriz-Req-Inf"

    # La cabecera es la primera fila y arranca en la A1: sin rótulos de
    # agrupación encima ni columnas en blanco por delante.
    encabezados = [nombre for nombre, _prefijo, _campo in CABECERA_MATRIZ]
    relleno = PatternFill("solid", fgColor="152342")
    for indice, titulo in enumerate(encabezados, start=1):     # la A es la 1
        celda = hoja.cell(row=FILA_CABECERA, column=indice, value=titulo)
        celda.font = Font(bold=True, color="FFFFFF", size=9)
        celda.fill = relleno
        celda.alignment = Alignment(wrap_text=True, vertical="center")

    for numero, fila in enumerate(filas, start=PRIMERA_FILA_DATOS):
        completa = dict(fila)
        for columna, calcular in RELLENO.items():
            completa.setdefault(columna, calcular(fila))
        for indice, titulo in enumerate(encabezados, start=1):
            hoja.cell(row=numero, column=indice, value=completa.get(titulo, ""))

    hoja.freeze_panes = f"A{PRIMERA_FILA_DATOS}"
    for indice in range(1, len(encabezados) + 1):
        hoja.column_dimensions[get_column_letter(indice)].width = 20
    libro.save(SALIDA)


if __name__ == "__main__":
    escribir(FILAS)
    oficios = {f["Referencia - Oficio FGE; Juzgado"] for f in FILAS}
    print(f"{SALIDA.name}: {len(FILAS)} filas -> {len(oficios)} oficios.")
    print("  Cabecera en la fila 1 (desde A1) y datos desde la fila 2.")
    print("  Las filas que comparten Referencia oficio se agrupan en un oficio, "
          "y cada una aporta una persona investigada.")

"""
Genera el archivo de datos de prueba para la carga masiva.

Crea `Matriz de prueba - 110 oficios.xlsx` con el FORMATO ESTABLECIDO (cabecera
en la fila 1, de la columna A a la Z, y los datos desde la fila 2) y 110 oficios
repartidos entre las dos instituciones.

Los datos se reparten a propósito para que el TABLERO se vea con contenido:

- **Fechas de recepción** repartidas por los últimos seis meses, con un grupo
  reciente en las dos últimas semanas y alguno de hoy mismo, para que los
  gráficos por día y por mes no salgan vacíos.
- **Responsables** con cargas distintas, de modo que el gráfico por responsable
  tenga barras de tamaños diferentes; unos pocos oficios entran sin responsable.
- **Estados** mezclados: mayoría finalizados, varios en proceso y algunos por
  asignar.
- **Tiempos de respuesta** variados, para que el promedio de días sea
  representativo.

- **Implicados**: cada oficio investiga a un número distinto de personas, de
  una sola a ocho, y de cada una se anota nombre, identificación, tipo de
  implicado y LCI. Como la matriz dedica una fila a cada persona, el archivo
  tiene bastantes más filas que oficios.

    python datos_de_prueba/generar_datos_prueba.py

No forma parte de la aplicación: es una utilidad para preparar una demostración
o para probar la carga masiva sin arriesgar datos reales.
"""
import random
import sys
from datetime import date, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from carga_masiva import (CABECERA_MATRIZ, FILA_CABECERA,   # noqa: E402
                          PRIMERA_FILA_DATOS)

SALIDA = Path(__file__).resolve().parent / "Matriz de prueba - 110 oficios.xlsx"
CANTIDAD_OFICIOS = 110

INSTITUCIONES = ["Superintendencia de Bancos", "Fiscalía General del Estado"]

# Los nombres se escriben como en la matriz real: inicial y apellido. Al
# importar, la aplicación los empareja con las cuentas del sistema.
#
# Son las cuentas con rol "usuario": los oficios se asignan a quien los
# tramita, no a quien administra el sistema. La cadena vacía es el oficio que
# llega sin responsable, que entra como "Por asignar".
#
# Cuántos oficios lleva cada uno (suman CANTIDAD_OFICIOS), para que el gráfico
# por responsable tenga barras claramente distintas.
CARGA_POR_USUARIO = {
    "C. Roman": 26,
    "J. Portero": 22,
    "J. Rosero": 18,
    "D. Franco": 15,
    "J. Galecio": 11,
    "L. Jarrin": 10,
    "": 8,                 # oficios que llegan sin responsable
}

TIPOS_ACCION = ["CERTIFICACIÓN", "RETENCIÓN", "INFORMACIÓN", "INMOVILIZACIÓN",
                "LEVANTAMIENTO DE MEDIDAS", "BLOQUEO Y RETENCIÓN",
                "RECTIFICACIÓN"]

DELITOS = [
    "TRAFICO ILÍCITO DE SUSTANCIAS CATALOGADAS SUJETAS A FISCALIZACIÓN",
    "LAVADO DE ACTIVOS", "COHECHO", "PECULADO", "ENRIQUECIMIENTO ILÍCITO",
    "DEFRAUDACIÓN TRIBUTARIA", "ESTAFA", "DESAPARICIÓN INVOLUNTARIA",
]

# Personas investigadas. La matriz dedica una fila a cada una, así que un
# oficio con cuatro implicados ocupa cuatro filas.
INVESTIGADOS = [
    "ORDOÑEZ VILLAGOMEZ DAVID MIGUEL", "ENOMENGA VARGAS ERICK LENIN",
    "BOLAÑOS RUBIO MARCO OLGER", "ACOSTA JEREZ DIANA CAROLINA",
    "MENDOZA SALAS LUIS ALBERTO", "PARRA NUÑEZ SOFIA ELENA",
    "CEVALLOS MORA JORGE ANDRÉS", "TAPIA LEÓN MARIA FERNANDA",
    "QUISPE ANDRADE PEDRO JOSÉ", "VILLACÍS ROJAS ANDREA PAOLA",
    "ZAMBRANO LOOR KEVIN DANIEL", "INTRIAGO CEDEÑO GLORIA ISABEL",
    "MACÍAS PALACIOS BYRON EDUARDO", "SUÁREZ VERA NATALIA CRISTINA",
    "CHÁVEZ ARIAS RAMIRO ANTONIO", "GUERRERO PINTO LUCÍA BELÉN",
    "COMERCIAL LOS ANDES S.A.", "IMPORTADORA DEL PACÍFICO CÍA. LTDA.",
    "AGRÍCOLA SANTA RITA S.A.", "TRANSPORTES DEL LITORAL CÍA. LTDA.",
]

# Cuántas personas investiga cada oficio. Se recorre en orden para que el
# archivo tenga de todo: oficios de una sola persona y otros de hasta ocho.
PATRON_INVESTIGADOS = [1, 1, 2, 1, 3, 1, 1, 4, 2, 1, 5, 1, 2, 1, 6,
                       3, 1, 2, 1, 8, 1, 1, 2, 7, 1, 3, 1, 1, 2, 1]

TIPOS_IDENTIFICACION = ["CED", "PAS", "RUC"]


def _identificacion(tipo, aleatorio):
    """Documento con el formato que exige cada tipo.

    La aplicación valida la cédula con 10 dígitos, el RUC con 13 y el pasaporte
    con letras y números, así que el archivo de prueba los genera ya válidos.
    """
    if tipo == "CED":
        return str(aleatorio.randint(10 ** 9, 10 ** 10 - 1))
    if tipo == "RUC":
        # Los RUC de persona natural son la cédula seguida de "001".
        return f"{aleatorio.randint(10 ** 9, 10 ** 10 - 1)}001"
    letras = "".join(aleatorio.choice("ABCDEFGHJKLMNPRSTUVWXYZ") for _ in range(2))
    return f"{letras}{aleatorio.randint(100000, 999999)}"
TIPOS_IMPLICADO = ["CLIENTE", "EX CLIENTE", "NO CLIENTE", "SIN IDENTIFICACION"]

MESES = ["ENE", "FEB", "MAR", "ABR", "MAY", "JUN",
         "JUL", "AGO", "SEP", "OCT", "NOV", "DIC"]


# Cómo se reparten las fechas de recepción, para que los gráficos del tablero
# tengan altibajos en vez de una línea plana.
#
# Días hacia atrás de los oficios recientes (0 = hoy). El gráfico por día cubre
# las dos últimas semanas: se repiten unos días y se saltan otros a propósito.
DIAS_RECIENTES = [0, 0, 0, 1, 1, 2, 2, 2, 2, 3, 3, 4, 4, 4, 5, 5, 6, 6,
                  6, 7, 7, 8, 8, 8, 9, 9, 10, 10, 11, 11, 12, 12, 13, 13, 13, 13]

# Cuántos oficios llegaron en cada mes anterior, del más reciente al más
# antiguo. Suman el resto de los oficios.
REPARTO_MESES = [20, 9, 16, 11, 18]


def _mes_atras(fecha, meses):
    """(año, mes) de la fecha indicada retrocediendo esa cantidad de meses."""
    mes = fecha.month - meses
    return fecha.year + (mes - 1) // 12, (mes - 1) % 12 + 1


def _fechas_recepcion(cantidad, hoy):
    """Fecha de recepción de cada oficio, repartida según los dos patrones."""
    fechas = [hoy - timedelta(days=dias) for dias in DIAS_RECIENTES]
    restantes = cantidad - len(fechas)
    for meses, cuantos in enumerate(REPARTO_MESES, start=1):
        anio, mes = _mes_atras(hoy, meses)
        for indice in range(min(cuantos, restantes)):
            # Días del 1 al 28: cualquier mes los tiene.
            fechas.append(date(anio, mes, 1 + indice * 27 // max(cuantos - 1, 1)))
        restantes -= cuantos
        if restantes <= 0:
            break
    # Si el reparto no llegó a cubrirlos todos, el resto va al mes más antiguo.
    while len(fechas) < cantidad:
        anio, mes = _mes_atras(hoy, len(REPARTO_MESES))
        fechas.append(date(anio, mes, 1 + len(fechas) % 28))
    return fechas[:cantidad]


def generar_filas():
    """Una fila por investigado. Algunos oficios repiten Referencia oficio para
    que la carga los agrupe y calcule la cantidad de investigados."""
    random.seed(2026)          # mismo archivo en cada ejecución
    hoy = date.today()
    fechas = _fechas_recepcion(CANTIDAD_OFICIOS, hoy)
    # Se reparten al azar para que la carga de cada persona no quede pegada a
    # un tramo de fechas concreto.
    responsables = [usuario
                    for usuario, cantidad in CARGA_POR_USUARIO.items()
                    for _ in range(cantidad)]
    random.shuffle(responsables)
    filas = []
    for numero in range(1, CANTIDAD_OFICIOS + 1):
        institucion = INSTITUCIONES[numero % 2]      # mitad y mitad
        sigla = "SB" if institucion.startswith("Super") else "FGE"

        recepcion = fechas[numero - 1]
        oficio = recepcion - timedelta(days=random.randint(1, 20))
        # Ninguna fecha puede ser futura: la aplicación las rechaza.
        asignacion = min(recepcion + timedelta(days=random.randint(0, 3)), hoy)

        usuario = responsables[numero - 1]
        # Lo recién llegado suele estar en proceso y lo antiguo ya respondido,
        # que es como se ve una carga real.
        antiguo = (hoy - recepcion).days > 20
        respondido = usuario and (antiguo or numero % 4 == 0)
        respuesta = None
        if respondido:
            # Tiempos de respuesta variados: de 1 a 30 días.
            respuesta = asignacion + timedelta(days=1 + (numero * 7) % 30)
            if respuesta > hoy:
                respuesta, respondido = None, False
        if not usuario:
            # Sin responsable no puede haber respuesta: la aplicación lo dejaría
            # "Por asignar" y le quitaría la fecha; se genera ya coherente.
            respuesta, respondido = None, False

        referencia_oficio = f"{sigla}-{recepcion.year}-{numero:04d}-OF"
        cuantos = PATRON_INVESTIGADOS[(numero - 1) % len(PATRON_INVESTIGADOS)]
        # Sin repetir persona dentro del mismo oficio.
        personas = random.sample(INVESTIGADOS, cuantos)
        for persona in personas:
            # Las empresas se identifican con RUC; las personas, con cualquiera
            # de los tres documentos.
            tipo_id = ("RUC" if persona.endswith(("S.A.", "LTDA."))
                       else random.choice(TIPOS_IDENTIFICACION))
            filas.append({
                "Institución del Estado": institucion,
                "Mes": MESES[recepcion.month - 1],
                "Fecha Asignación": asignacion if usuario else None,
                "Usuario": usuario,
                "Prioridad": random.choice(["Alta", "Media", "Baja"]),
                "Fecha Emisión": recepcion,
                "Referencia": f"{str(recepcion.year)[2:]}-{numero:04d}-UDC",
                "Medio Repuesta": random.choice(["Electrónico", "Físico"]),
                "Fecha Envío": respuesta,
                "Estado": "Finalizado" if respondido else (
                    "En proceso" if usuario else "Por asignar"),
                "Días": (respuesta - asignacion).days if respuesta else "",
                "Canal Recepc": random.choice(["Proveedor", "Ventanilla",
                                               "Correo"]),
                "Fecha Circular": oficio,
                "Apellidos, Nombres - Razón Social": persona,
                "TiPASo Id CED; PAS; RUCUC": tipo_id,
                "Identificación Ced; Pas; RUC": _identificacion(tipo_id, random),
                "Referencia - Oficio FGE; Juzgado": referencia_oficio,
                "Número Expediente Fiscal": "-",
                "Referencia - Circular Superintendencia Bancos":
                    f"SB-SG-{recepcion.year}-{random.randint(10000, 99999)}-C",
                "Delito": DELITOS[numero % len(DELITOS)],
                "Tipo de Accion": TIPOS_ACCION[numero % len(TIPOS_ACCION)],
                "Observación": random.choice(
                    ["", "", "Atendido dentro del plazo",
                     "Requiere seguimiento", "Se remitió por correo"]),
                "Tipo de Implicado": random.choice(TIPOS_IMPLICADO),
                "LCI - SI o NO": random.choice(["SI", "NO", "NO"]),
                "Fecha - Solicitud": recepcion + timedelta(days=1),
                "Ref Solic- No. LCI-202X-000": f"LCI-{recepcion.year}-"
                                               f"{random.randint(1, 99):03d}",
            })
    return filas


def escribir(filas):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Matriz-Req-Inf"

    # La cabecera es la primera fila y arranca en la A1: sin rótulos de
    # agrupación encima ni columnas en blanco por delante.
    encabezados = [nombre for nombre, _prefijo, _campo in CABECERA_MATRIZ]
    relleno = PatternFill("solid", fgColor="152342")
    for indice, titulo in enumerate(encabezados, start=1):   # la A es la 1
        celda = hoja.cell(row=FILA_CABECERA, column=indice, value=titulo)
        celda.font = Font(bold=True, color="FFFFFF", size=9)
        celda.fill = relleno
        celda.alignment = Alignment(wrap_text=True, vertical="center")

    for numero, fila in enumerate(filas, start=PRIMERA_FILA_DATOS):
        for indice, titulo in enumerate(encabezados, start=1):
            hoja.cell(row=numero, column=indice, value=fila.get(titulo, ""))

    hoja.freeze_panes = f"A{PRIMERA_FILA_DATOS}"
    for indice in range(1, len(encabezados) + 1):
        hoja.column_dimensions[hoja.cell(row=FILA_CABECERA,
                                         column=indice).column_letter].width = 18
    libro.save(SALIDA)


def _resumen(filas):
    """Cómo quedó repartido lo generado (para verlo al ejecutar el script)."""
    from collections import Counter
    oficios = {}
    for fila in filas:
        oficios.setdefault(fila["Referencia - Oficio FGE; Juzgado"], fila)
    estados = Counter(f["Estado"] for f in oficios.values())
    por_oficio = Counter()
    for fila in filas:
        por_oficio[fila["Referencia - Oficio FGE; Juzgado"]] += 1
    reparto = Counter(por_oficio.values())
    usuarios = Counter(f["Usuario"] or "(sin responsable)"
                       for f in oficios.values())
    entidades = Counter(f["Institución del Estado"] for f in oficios.values())
    return oficios, estados, usuarios, entidades, reparto


if __name__ == "__main__":
    filas = generar_filas()
    escribir(filas)
    oficios, estados, usuarios, entidades, reparto = _resumen(filas)
    print(f"{SALIDA.name}: {len(filas)} filas -> {len(oficios)} oficios "
          f"(las filas que comparten Referencia oficio se agrupan).")
    print("  Estados:      " + ", ".join(f"{k}: {v}" for k, v in estados.items()))
    print("  Instituciones:" + ", ".join(f" {k}: {v}" for k, v in entidades.items()))
    print("  Responsables: " + ", ".join(f"{k}: {v}"
                                         for k, v in usuarios.most_common()))
    print("  Implicados:   " + ", ".join(
        f"{cuantos} implicado(s): {cuantos_oficios} oficio(s)"
        for cuantos, cuantos_oficios in sorted(reparto.items())))

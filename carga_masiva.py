"""
Carga masiva de oficios (importación) desde un archivo .xlsx o .csv.

Formato establecido: EL DE LA EXPORTACIÓN, SIN LA REFERENCIA UDC
-----------------------------------------------------------------
El archivo que se importa tiene las columnas que produce *Exportar oficios*, en
su mismo orden (ver `almacen_oficios.COLUMNAS_EXPORTACION` y
`COLUMNAS_IMPLICADO`), **menos la Referencia UDC**: la numera el sistema al
importar, con la nomenclatura de la institución de cada oficio, así que pedirla
solo daría pie a escribir una que no se va a usar. La cabecera ocupa la fila 1,
desde la celda A1, y los datos empiezan en la fila 2. Así hay un solo formato
que mantener y lo que sale del sistema se parece a lo que entra.

Como en la exportación, **cada fila es una persona investigada**: las filas que
comparten la misma *Referencia oficio* son el mismo oficio, y de ellas sale su
detalle de implicados. Los datos del oficio se repiten en cada una de sus filas
y tienen que coincidir; si no, se avisa.

Columnas que sí están en el archivo pero cuyo contenido **rellena el sistema**
y se ignora al importar: Documento del oficio, Respuesta en PDF, Registrado por,
Fecha de registro y Origen.

`escribir_plantilla()` escribe un .xlsx con este mismo formato, y es lo que usan
los archivos de ejemplo de `datos_de_prueba/`: el módulo que lee el formato es
el que lo escribe.

Todo o nada
-----------
Antes de guardar nada se valida el archivo ENTERO con las mismas reglas que
aplica el sistema al registrar un oficio a mano (`almacen_oficios`): fechas
coherentes, estado acorde al responsable, responsable existente, tipo de acción
del catálogo, implicados con su identificación bien formada, referencias sin
repetir… Si una sola fila falla, **no se importa nada**: se muestran todas las
filas con error, con su línea del archivo y su motivo, para corregirlas de una
pasada y volver a cargar.

Diferencias con el alta manual, por lo que un archivo no puede aportar:
  - No se exige el documento del oficio ni la respuesta en PDF; se adjuntan
    después desde la pestaña Oficios. Un oficio ya finalizado sí exige sus
    fechas de asignación y de respuesta.
  - El responsable se indica por su **nombre de cuenta** (columna «Usuario
    responsable»), que debe existir en el sistema.
"""
import csv
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import almacen_oficios

# --- Formato del archivo -----------------------------------------------------
# La cabecera es la primera fila y los datos empiezan en la segunda, igual que
# en el archivo que produce la exportación.
FILA_CABECERA = 1
PRIMERA_FILA_DATOS = FILA_CABECERA + 1

# Columnas del implicado: se les pone prefijo para que no choquen con las del
# oficio (por ejemplo "identificacion") al leer la fila.
PREFIJO_IMPLICADO = "implicado_"

# La Referencia UDC NO forma parte del archivo de carga: la numera el sistema al
# importar, con la nomenclatura de la institución de cada oficio, así que pedirla
# solo daría pie a escribir una que no se va a usar.
CAMPO_EXCLUIDO = "referencia"

# El formato se DERIVA de la exportación, no se copia: si allí se añade una
# columna, aquí aparece sola y sigue habiendo un único formato.
COLUMNAS = (
    [(clave, titulo, "oficio")
     for clave, titulo in almacen_oficios.COLUMNAS_EXPORTACION.items()
     if clave != CAMPO_EXCLUIDO]
    + [(PREFIJO_IMPLICADO + clave, titulo, "implicado")
       for clave, titulo in almacen_oficios.COLUMNAS_IMPLICADO.items()]
)
CABECERA = [titulo for _clave, titulo, _ambito in COLUMNAS]
TITULO_EXCLUIDO = almacen_oficios.COLUMNAS_EXPORTACION[CAMPO_EXCLUIDO]

# Columnas que sí están en el archivo pero cuyo contenido rellena el sistema.
CAMPOS_ASIGNADOS = {
    "archivo_oficio": "se adjunta después",
    "archivo_respuesta": "se adjunta después",
    "registrado_por": "es quien importa el archivo",
    "fecha_registro": "es la de la importación",
    "origen": "queda como «carga masiva»",
}

# Columnas cuyo valor es una fecha (AAAA-MM-DD en la exportación).
CAMPOS_FECHA = {"fecha_oficio", "fecha_recepcion",
                "fecha_asignacion", "fecha_respuesta"}

_ACENTOS = str.maketrans("áéíóúàèìòùäëïöüâêîôûÁÉÍÓÚÄËÏÖÜÂÊÎÔÛ",
                         "aeiouaeiouaeiouaeiouAEIOUAEIOUAEIOU")


def normalizar(texto) -> str:
    """Texto en minúsculas, sin tildes y con los espacios colapsados."""
    if texto is None:
        return ""
    return " ".join(str(texto).translate(_ACENTOS).lower().split())


def _a_fecha(valor) -> str:
    """Convierte a 'AAAA-MM-DD'. Devuelve '' si la celda está vacía.

    Acepta lo que Excel entrega como fecha real y también las fechas escritas a
    mano en los formatos habituales, incluido el día/mes/año que se usa aquí.
    """
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        return valor.date().isoformat()
    if isinstance(valor, date):
        return valor.isoformat()
    texto = str(valor).strip()
    if not texto or texto in {"-", "--", "N/A", "n/a"}:
        return ""
    # Excel a veces deja la hora pegada a la fecha.
    texto = texto.split(" ")[0].replace("/", "-").replace(".", "-")
    for formato in ("%Y-%m-%d", "%d-%m-%Y", "%m-%d-%Y", "%d-%m-%y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(texto, formato).date().isoformat()
        except ValueError:
            continue
    raise ValueError(f"No se reconoce la fecha «{valor}»")


def _a_texto(valor) -> str:
    """Contenido de una celda como texto limpio. Los guiones sueltos que se usan
    para 'sin dato' se tratan como vacío."""
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)          # 1650110974001.0 -> 1650110974001
    texto = " ".join(str(valor).split())
    return "" if texto in {"-", "--", "N/A", "n/a"} else texto


# --- Validación de la cabecera -----------------------------------------------
def _letra_columna(posicion: int) -> str:
    """Letra de la columna de Excel para la posición indicada (0 -> 'A')."""
    numero = posicion + 1
    letras = ""
    while numero:
        numero, resto = divmod(numero - 1, 26)
        letras = chr(65 + resto) + letras
    return letras


PRIMERA_COLUMNA = _letra_columna(0)
ULTIMA_COLUMNA = _letra_columna(len(COLUMNAS) - 1)


def _recortar(celdas: List) -> Tuple[List[str], List[str]]:
    """Encabezados de la fila, sin las columnas vacías del final.

    Devuelve (normalizados, tal como vienen): los normalizados sirven para
    comparar y los originales para los mensajes, que así muestran el texto
    exacto que tiene el archivo. Solo se descartan las columnas sobrantes del
    final, que Excel arrastra en blanco; las del principio NO, porque la
    cabecera tiene que empezar en la A.
    """
    titulos = [normalizar(c) for c in celdas]
    originales = [" ".join(str(c).split()) if c is not None else "" for c in celdas]
    fin = len(titulos)
    while fin > 0 and not titulos[fin - 1]:
        fin -= 1
    return titulos[:fin], originales[:fin]


def _error_formato(detalle: str) -> ValueError:
    """Rechazo del archivo por no tener el formato establecido."""
    return ValueError(
        f"El archivo no tiene el formato establecido: {detalle}\n\n"
        f"Debe tener las mismas columnas que la exportación de oficios —salvo "
        f"la «{TITULO_EXCLUIDO}», que no se toma del archivo— en su mismo "
        f"orden: la cabecera en la fila {FILA_CABECERA}, de la columna "
        f"{PRIMERA_COLUMNA} a la {ULTIMA_COLUMNA}, con las {len(COLUMNAS)} "
        f"columnas, y los datos a partir de la fila {PRIMERA_FILA_DATOS}.\n\n"
        "Use el archivo de ejemplo de la carga masiva como plantilla."
    )


def validar_cabecera(celdas: List) -> None:
    """Comprueba que la fila sea la cabecera del formato establecido.

    El orden de las columnas IMPORTA: se exige la secuencia completa, empezando
    en la primera celda. Solo se toleran diferencias de redacción (mayúsculas,
    tildes y espacios de más). Si el archivo no cumple, lanza un ValueError
    explicando qué falla.
    """
    titulos, originales = _recortar(celdas)
    if not titulos:
        raise _error_formato(f"la fila {FILA_CABECERA} está vacía.")
    if not titulos[0]:
        vacias = 0
        while vacias < len(titulos) and not titulos[vacias]:
            vacias += 1
        raise _error_formato(
            f"la cabecera no empieza en la columna {PRIMERA_COLUMNA} (hay "
            f"{vacias} columna(s) en blanco por delante).")
    if titulos[0] == normalizar(TITULO_EXCLUIDO):
        # Caso típico de quien parte de un archivo exportado: la exportación
        # lleva la Referencia UDC delante y la carga no la usa.
        raise _error_formato(
            f"sobra la columna {PRIMERA_COLUMNA} «{TITULO_EXCLUIDO}». La "
            f"numera el sistema al importar, con la nomenclatura de la "
            f"institución de cada oficio, así que no se toma del archivo: "
            f"elimine esa columna.")

    problemas = []
    for posicion, (_clave, titulo, _ambito) in enumerate(COLUMNAS):
        letra = _letra_columna(posicion)
        if posicion >= len(titulos):
            problemas.append(f"falta la columna {letra} «{titulo}»")
        elif titulos[posicion] != normalizar(titulo):
            encontrado = originales[posicion][:40] or "(vacía)"
            problemas.append(
                f"la columna {letra} debería ser «{titulo}» y contiene "
                f"«{encontrado}»")
    sobran = len(titulos) - len(COLUMNAS)
    if sobran > 0:
        problemas.append(
            f"hay {sobran} columna(s) de más después de la {ULTIMA_COLUMNA}")

    if problemas:
        detalle = "\n".join(f"  · {p}" for p in problemas[:8])
        if len(problemas) > 8:
            detalle += f"\n  · … y {len(problemas) - 8} diferencia(s) más"
        raise _error_formato(f"las columnas no coinciden.\n\n{detalle}")


# --- Lectura del archivo -----------------------------------------------------
def _leer_xlsx(ruta: Path) -> List[List]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ValueError(
            "Para leer archivos de Excel hace falta la librería openpyxl:\n\n"
            "    pip install openpyxl\n\n"
            "También puede guardar el archivo como CSV y cargarlo así."
        )
    libro = load_workbook(ruta, data_only=True, read_only=True)
    hoja = libro.active
    filas = [list(f) for f in hoja.iter_rows(values_only=True)]
    libro.close()
    if not filas:
        raise _error_formato("el archivo está vacío.")
    return filas


def _leer_csv(ruta: Path) -> List[List]:
    """Lee un CSV detectando el separador (barra vertical, punto y coma o coma).

    La cabecera es la PRIMERA línea, igual que en la hoja de Excel. Las líneas
    se devuelven todas, sin descartar las vacías, para que el número de fila que
    se informe en los errores sea el que de verdad tiene el archivo.
    """
    with ruta.open("r", encoding="utf-8-sig", newline="") as archivo:
        muestra = archivo.read(8192)
        archivo.seek(0)
        separador = max(almacen_oficios.SEPARADOR_CSV + ";,\t", key=muestra.count)
        filas = [f for f in csv.reader(archivo, delimiter=separador)]
    if not filas:
        raise _error_formato("el archivo está vacío.")
    return filas


def error_de_fila(numero, codigo_oficio, motivo) -> Dict:
    """Describe una fila que no se puede importar.

    Se devuelve en piezas —número de fila, Referencia oficio y motivo— y no
    como una frase ya armada, para que la aplicación pueda listarlas en una
    tabla y quien cargue el archivo sepa exactamente qué línea corregir.
    """
    motivo = str(motivo)
    return {"fila": str(numero), "codigo_oficio": (codigo_oficio or "").strip(),
            # Los motivos vienen de sitios distintos y unos empiezan en
            # minúscula; en una columna de una tabla desentonan.
            "motivo": motivo[:1].upper() + motivo[1:]}


def ordenar_errores(errores: List[Dict]) -> List[Dict]:
    """Los ordena por línea del archivo.

    Quien corrige el archivo lo recorre de arriba abajo, no por el momento en
    que se detectó cada problema.
    """
    def _primera_linea(error):
        primera = str(error.get("fila", "")).split(",")[0].strip()
        return int(primera) if primera.isdigit() else 0
    return sorted(errores, key=_primera_linea)


def etiqueta_filas(fila: Dict) -> str:
    """Número(s) de línea del archivo que componen el oficio ('11' o '11, 12')."""
    numeros = fila.get("_filas") or [fila.get("_fila", "?")]
    return ", ".join(str(n) for n in numeros)


def leer_archivo(ruta) -> Tuple[List[Dict], List[Dict]]:
    """Lee el archivo y devuelve (filas, errores).

    Cada fila es un diccionario con los campos ya normalizados y una clave
    `_fila` con su número dentro del archivo, para poder señalar los errores.
    Los errores son los que devuelve `error_de_fila`.
    """
    ruta = Path(ruta)
    if not ruta.exists():
        raise ValueError("No se encontró el archivo seleccionado.")
    if ruta.suffix.lower() in (".xlsx", ".xlsm"):
        crudas = _leer_xlsx(ruta)
    elif ruta.suffix.lower() == ".csv":
        crudas = _leer_csv(ruta)
    else:
        raise ValueError("El archivo debe ser una hoja de Excel (.xlsx) o un CSV.")

    # El formato es fijo: se comprueba la cabecera antes de leer un solo dato.
    validar_cabecera(crudas[0])

    filas, errores = [], []
    for desplazamiento, cruda in enumerate(crudas[1:], start=1):
        numero = FILA_CABECERA + desplazamiento
        datos = {"_fila": numero}
        problema = None
        for posicion, (clave, _titulo, _ambito) in enumerate(COLUMNAS):
            valor = cruda[posicion] if posicion < len(cruda) else None
            if clave in CAMPOS_FECHA:
                try:
                    datos[clave] = _a_fecha(valor)
                except ValueError as error:
                    problema = str(error)
                    datos[clave] = ""
            else:
                datos[clave] = _a_texto(valor)
        if not any(datos[clave] for clave, _t, _a in COLUMNAS):
            continue                      # fila en blanco
        if problema:
            errores.append(error_de_fila(numero, datos.get("codigo_oficio"),
                                         problema))
            continue
        filas.append(datos)
    return filas, errores


# --- Escritura: la plantilla de la carga -------------------------------------
def filas_plantilla(oficios: List[Dict]) -> List[List[str]]:
    """Las filas del archivo de carga para esos oficios: una por implicado.

    Es la exportación sin la Referencia UDC. Sirve para escribir plantillas y
    archivos de ejemplo con el formato exacto que la carga espera leer, sin que
    nadie tenga que reproducirlo a mano.
    """
    filas = []
    for oficio in oficios:
        datos = [_a_texto(oficio.get(clave, "")) for clave, _t, ambito in COLUMNAS
                 if ambito == "oficio"]
        implicados = oficio.get("implicados") or [{}]
        for implicado in implicados:
            filas.append(datos + [
                _a_texto(implicado.get(clave[len(PREFIJO_IMPLICADO):], ""))
                for clave, _t, ambito in COLUMNAS if ambito == "implicado"])
    return filas


def escribir_plantilla(oficios: List[Dict], ruta_destino: str) -> None:
    """Escribe un .xlsx con el formato de la carga masiva."""
    almacen_oficios.escribir_xlsx(CABECERA, filas_plantilla(oficios),
                                  ruta_destino, hoja_titulo="Oficios")


# --- Agrupación: una fila por implicado --------------------------------------
def _implicado_de(fila: Dict) -> Optional[Dict]:
    """Datos del implicado de una fila, o None si la fila no trae ninguno.

    No se valida aquí: de eso se encarga el almacén con la misma regla que usa
    el alta manual, para que el archivo y el formulario exijan lo mismo.
    """
    datos = {clave[len(PREFIJO_IMPLICADO):]: fila.get(clave, "")
             for clave, _titulo, ambito in COLUMNAS if ambito == "implicado"}
    return datos if any(datos.values()) else None


# Columnas del oficio que se repiten en cada una de sus filas y tienen que
# decir lo mismo. Se excluyen las que rellena el sistema (su contenido se
# ignora) y la cantidad de investigados, que la cuenta el propio detalle.
def _campos_comparables() -> List[Tuple[str, str]]:
    return [(clave, titulo) for clave, titulo, ambito in COLUMNAS
            if ambito == "oficio" and clave not in CAMPOS_ASIGNADOS
            and clave != "cantidad_investigados"]


def agrupar_por_referencia(filas: List[Dict]) -> Tuple[List[Dict], List[Dict]]:
    """Une las filas que comparten Referencia oficio en un solo oficio.

    Cada fila aporta una persona investigada, así que un oficio con cuatro
    implicados ocupa cuatro filas con los mismos datos de oficio repetidos. Si
    esas copias no coinciden, el archivo se contradice y se informa: se devuelve
    (oficios, errores).
    """
    agrupados: Dict[str, Dict] = {}
    orden: List[str] = []
    errores: List[Dict] = []
    comparables = _campos_comparables()

    for fila in filas:
        codigo = (fila.get("codigo_oficio") or "").strip()
        # Sin Referencia oficio no hay con qué agrupar: la fila va sola y el
        # almacén la rechazará por no tenerla.
        clave = codigo.upper() or f"__fila_{fila['_fila']}"
        implicado = _implicado_de(fila)
        if clave not in agrupados:
            copia = dict(fila)
            copia["implicados"] = [implicado] if implicado else []
            # Todas las líneas del archivo que forman el oficio: si el oficio
            # no se puede importar hay que poder señalarlas todas.
            copia["_filas"] = [fila["_fila"]]
            agrupados[clave] = copia
            orden.append(clave)
            continue

        oficio = agrupados[clave]
        oficio["_filas"].append(fila["_fila"])
        if implicado:
            oficio["implicados"].append(implicado)
        discrepan = [titulo for campo, titulo in comparables
                     if normalizar(fila.get(campo)) != normalizar(oficio.get(campo))]
        if discrepan:
            errores.append(error_de_fila(
                fila["_fila"], codigo,
                f"esta línea contradice a la línea {oficio['_filas'][0]} del "
                f"mismo oficio en: {', '.join(discrepan[:4])}"
                + (" …" if len(discrepan) > 4 else "") + "."))
    return [agrupados[c] for c in orden], errores


# --- Preparación -------------------------------------------------------------
def preparar(ruta, actor: str = "", actor_rol: str = "") -> Dict:
    """Deja los oficios listos para importar y resume lo que se va a hacer.

    Valida el archivo entero por adelantado —con las reglas del propio
    almacén—, de modo que la vista previa dice exactamente lo que va a pasar y
    la importación solo se ofrece si no queda ninguna fila con error.
    """
    filas, errores = leer_archivo(ruta)
    oficios, incoherencias = agrupar_por_referencia(filas)
    errores += incoherencias
    # Los oficios que ya se contradicen entre sus propias líneas no se vuelven
    # a validar: bastante ruido hay con señalar la contradicción.
    contradictorios = {e["fila"] for e in incoherencias}
    revisables = [o for o in oficios
                  if not contradictorios & {str(n) for n in o["_filas"]}]
    errores += almacen_oficios.validar_importacion(revisables, actor, actor_rol)

    return {
        "filas": oficios,
        "errores": ordenar_errores(errores),
        "campos_asignados": [almacen_oficios.COLUMNAS_EXPORTACION[clave]
                             for clave in CAMPOS_ASIGNADOS],
    }
